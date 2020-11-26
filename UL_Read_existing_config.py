
#Author: Ashritha
#16-Nov-2018: Create - Basic version to read variables from stored files
#16-Jan-2019: Multiple updates - new variables captured as per need
#23-Sep-2019: Modified for UL
#15-Oct-2019: Modified for reading interfaces 
#16-Oct-2019: Modified for reading RFP ref from assessment sheet based on device name
#			  Modified for reading Netmask ordered from Master Tracker
#			  Modified for routing information and class-maps and summary sheet for each router
#			  Modified output format red to yellow, freeze first row, format first row in all sheets, deleted default sheet
#26-Nov-2019: Modified to include HSRP priority
#13-Mar-2020:
#07-Apr-2020:
#25-May-2020: 
#23-jun-2020: Added HSRP to summary sheet.
#22-Jul-2020: Added checks for SDWAN. Added checks for WAN resiliency. Added checks for WAN checks as agreed checklist with Dave and team

#Pending: ZBFW attach points

import os,re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
from openpyxl.cell import Cell
from datetime import datetime
import pandas as pd
import time
import numpy as np
import json
import smartsheet, logging
import SubnetTree, itertools
from pathlib import Path



#show ip route parsing
# Path to directory with routing table files.# Each routing table MUST be in separate .txt file.
RT_DIRECTORY = "./routing_tables"

# RegEx template string for IPv4 address matching.
REGEXP_IPv4_STR = (
      '((25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.'
    + '(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.'
    + '(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.'
    + '(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?))'
)
# IPv4 CIDR notation matching in user input.
REGEXP_INPUT_IPv4 = re.compile("^" + REGEXP_IPv4_STR + "(\/\d\d?)?$")
# Local and Connected route strings matching.
REGEXP_ROUTE_LOCAL_CONNECTED = re.compile(
     '^(?P<routeType>[L|C])\s+'
    + '((?P<ipaddress>\d\d?\d?\.\d\d?\d?\.\d\d?\d?\.\d\d?\d?)'
    + '\s?'
    + '(?P<maskOrPrefixLength>(\/\d\d?)?'
    + '|(\d\d?\d?\.\d\d?\d?\.\d\d?\d?\.\d\d?\d?)?))'
    + '\ is\ directly\ connected\,\ '
    + '(?P<interface>\S+)',
    re.MULTILINE
)
# Static and dynamic route strings matching.
REGEXP_ROUTE = re.compile(
      '^(\S\S?\*?\s?\S?\S?)'
    + '\s+'
    + '((?P<subnet>\d\d?\d?\.\d\d?\d?\.\d\d?\d?\.\d\d?\d?)'
    + '\s?'
    + '(?P<maskOrPrefixLength>(\/\d\d?)?'
    +'|(\d\d?\d?\.\d\d?\d?\.\d\d?\d?\.\d\d?\d?)?))'
    + '\s*'
    + '(?P<viaPortion>(?:\n?\s+(\[\d\d?\d?\/\d+\])\s+'
    + 'via\s+(\d\d?\d?\.\d\d?\d?\.\d\d?\d?\.\d\d?\d?)(.*)\n?)+)',
    re.MULTILINE
)
# Route string VIA portion matching.
REGEXP_VIA_PORTION = re.compile(
    '.*via\s+(\d\d?\d?\.\d\d?\d?\.\d\d?\d?\.\d\d?\d?).*'
)
# Store for 'router' objects generated from input routing table files. # Each file is represented by single 'router' object.# Router is referenced by Router ID (RID).# RID is filename by default.# Format:## ROUTERS = {#     'RID1': {'routing_table': {}, 'interface_list': ()},#     'RID_N': {'routing_table': {}, 'interface_list': ()},# }# 
ROUTERS = {}
# Global search tree for Interface IP address to Router ID (RID) resolving.# Stores Interface IP addresses as keys.# Returns (RID, interfaceID) list.# Interface IP addresses SHOULD be globally unique across inspected topology.
GLOBAL_INTERFACE_TREE = SubnetTree.SubnetTree()

def parse_show_ip_route_ios_like(raw_routing_table):
	"""
	Parser for routing table text output.
	Compatible with both Cisco IOS(IOS-XE) 'show ip route' 
	and Cisco ASA 'show route' output format.
	Processes input text file and writes into Python data structures.
	Builds internal SubnetTree search tree in 'route_tree'.
	Generates local interface list for router in 'interface_list'
	Returns 'router' dictionary object with parsed data.
	"""
	router = {}
	route_tree = SubnetTree.SubnetTree()
	interface_list = []
	dict_routes = {}
	# Parse Local and Connected route strings in text.
	for raw_route_string in REGEXP_ROUTE_LOCAL_CONNECTED.finditer(raw_routing_table):
		#print ("RRS LC", raw_route_string)
		subnet = (
			raw_route_string.group('ipaddress') 
			+ convert_netmask_to_prefix_length(
			raw_route_string.group('maskOrPrefixLength')
			)
		)
		
		interface = raw_route_string.group('interface')
		#print (interface)
		route_tree[subnet] = ((interface,), raw_route_string.group(0))
		dict_routes[subnet] = ((interface,), raw_route_string.group(0))
		if raw_route_string.group('routeType') == 'L':
			interface_list.append((interface, subnet,))
	if not interface_list:
		print('Failed to find routing table entries in given output')
		return None# parse static and dynamic route strings in text
	for raw_route_string in REGEXP_ROUTE.finditer(raw_routing_table):
			#print ("RRS RT", raw_route_string)
			subnet = ( raw_route_string.group('subnet') 
					+ convert_netmask_to_prefix_length(
					raw_route_string.group('maskOrPrefixLength')))
			via_portion =  raw_route_string.group('viaPortion')
			next_hops= []
			if via_portion.count('via') > 1:
				for line in via_portion.splitlines():
					if line:
						next_hops.append(REGEXP_VIA_PORTION.match(line).group(1))
			else:
				next_hops.append(REGEXP_VIA_PORTION.match(via_portion).group(1))
			route_tree[subnet] = (next_hops, raw_route_string.group(0))
			dict_routes[subnet] = (next_hops, raw_route_string.group(0))
			#print ("Next hops ", next_hops, "\n RRS grp ",raw_route_string.group(0))
	#print ("RT: ",route_tree)
	#print ("interface_list:", interface_list)
	#print ("dictionary: ", dict_routes)
	router = {
		'routing_table': dict_routes, #used to be route_tree
		'interface_list': interface_list,
		'subnet_tree_routing_table': route_tree
	}
	return router

def parse_text_routing_table(raw_routing_table):
	"""
	Parser functions wrapper.
	Add additional parsers for alternative routing table syntaxes here.
	"""
	router = parse_show_ip_route_ios_like(raw_routing_table)
	if router:
		return router
	    
def do_parse_routing_table(show_ip_route_dict):
	new_routers = {}
	for label in show_ip_route_dict["Host Name"]:
		new_router = parse_text_routing_table(show_ip_route_dict["show ip route"])
		router_id = label
		if new_router:
			new_routers[router_id] = new_router
			if new_router['interface_list']:
				for iface, addr in new_router['interface_list']:
					GLOBAL_INTERFACE_TREE[addr]= (router_id, iface,)
			else:
				print ('Failed to parse routiung table ' + label)
	return new_routers

def do_parse_dict(sh_ip_rt_from_dict,router_id):
	"""
	Go through specified dictionary element and parse.
	Generate router objects based on parse result if any.
	Populate new_routers with those router objects.
	Default key for each router object is FILENAME.
	Return new_router.
	"""
	#new_routers = {}
	if not sh_ip_rt_from_dict:
		print ("show ip route not parseable")
		return {}
	new_router = parse_text_routing_table(sh_ip_rt_from_dict)
	
	if new_router:
		#new_routers[router_id] = new_router
		if new_router['interface_list']:
			for iface, addr in new_router['interface_list']:
				GLOBAL_INTERFACE_TREE[addr]= (router_id, iface,)
		else:
			print ('Failed to parse ' + hn)
	else:
		print ("\nEmpty show ip route") # in{} sec".format("{:.3f}".format(time() - start_time)))
	
	return new_router

def compare_routes_of_wan(rtr_dict,op_dir,region_site):
	combo_op = {}
	wb = Workbook()
	ws_summ = wb.create_sheet("Summary sheet")
	row_summ = ws_summ.max_row
	ws_summ.cell(row_summ,1,"Country")
	ws_summ.cell(row_summ,2,"Work Location")
	ws_summ.cell(row_summ,3,"Hostname")
	ws_summ.cell(row_summ,4,"Neighbor")
	ws_summ.cell(row_summ,5,"OSPF DIO on Hostname")
	ws_summ.cell(row_summ,6,"OSPF DIO on Neighbor")
	nei_num = len(rtr_dict.keys())
	if nei_num < 1:
		return nei_num, combo_op
	if nei_num == 1:
		for i in rtr_dict.keys():
			combo_op[i] = {}
			ws = wb.create_sheet("Single router site")
			row = ws.max_row
			ws.cell(row,1,"Network")
			ws.cell(row,2,"Route in "+i)
			ws.cell(row,3,"Route in neighbor")
			write_cell(ws_summ,row_summ+1,1,region_site.split(":")[1])
			write_cell(ws_summ,row_summ+1,2,region_site.split(":")[2])
			write_cell(ws_summ,row_summ+1,3,i)

			
			if "sir_parsed" in rtr_dict[i].keys():
				if "dict" in str(type (rtr_dict[i]["sir_parsed"])):
					if "subnet_tree_routing_table" in rtr_dict[i]["sir_parsed"].keys():
						for network in rtr_dict[i]["sir_parsed"]['routing_table']:
							#print ("Checking network for one device", network)
							combo_op[i]["dummy"] = {}
							combo_op[i]["dummy"]["Matched"] = []
							combo_op[i]["dummy"]["Only in Rtr1"] = []
							combo_op[i]["dummy"]["Only in Rtr2"] = []
							if "OSPF DIO metrics" in rtr_dict[i].keys():
								combo_op[i]["dummy"]["OSPF DIO metrics1"] = rtr_dict[i]["OSPF DIO metrics"]
								if 'list' in str(type(rtr_dict[i]["OSPF DIO metrics"])):
									write_cell(ws_summ,row_summ+1,5,list_to_str(rtr_dict[i]["OSPF DIO metrics"]))
								else:
									write_cell(ws_summ,row_summ+1,5,str(rtr_dict[i]["OSPF DIO metrics"]))

							row = ws.max_row
							write_cell(ws, row+1, 1, network)
							if network in rtr_dict[i]["sir_parsed"]["subnet_tree_routing_table"]:
								combo_op[i]["dummy"]["Only in Rtr1"].append(network)
								next_hop1, raw_route_string1 = rtr_dict[i]["sir_parsed"]["subnet_tree_routing_table"][network]
								write_cell(ws, row+1, 2, str(raw_route_string1))
								write_cell(ws, row+1, 3, "No Neighbor")
								#print ("route only in RTR1")
		wb.remove(wb['Sheet'])
		#freeze_panes (wb,"B1")
		wb.save(os.path.join(op_dir,region_site+".xlsx"))
		return nei_num, combo_op
	nei_grp_num = 0

	for x, y in itertools.combinations(rtr_dict.keys(),2):
		nei_grp_num+=1
		combo_op[x] = {}
		combo_op[x][y] = {}
		combo_op[x][y]["Matched"] = []
		combo_op[x][y]["Only in Rtr1"] = []
		combo_op[x][y]["Only in Rtr2"] = []
		ws = wb.create_sheet("Nbr-group"+str(nei_grp_num))
		row = ws.max_row
		ws.cell(row,1,"Network")
		ws.cell(row,2,"Route in "+x)
		ws.cell(row,3,"Route in "+y)
		write_cell(ws_summ,row_summ+1,1,region_site.split(":")[1])
		write_cell(ws_summ,row_summ+1,2,region_site.split(":")[2])
		write_cell(ws_summ,row_summ+1,3,x)
		write_cell(ws_summ,row_summ+1,4,y)
		if "OSPF DIO metrics" in rtr_dict[x].keys():
			combo_op[x][y]["OSPF DIO metrics1"] = rtr_dict[x]["OSPF DIO metrics"]
			if 'list' in str(type(rtr_dict[x]["OSPF DIO metrics"])):
				write_cell(ws_summ,row_summ+1,5,list_to_str(rtr_dict[x]["OSPF DIO metrics"]))
			else:
				write_cell(ws_summ,row_summ+1,5,str(rtr_dict[x]["OSPF DIO metrics"]))
		if "OSPF DIO metrics" in rtr_dict[y].keys():
			combo_op[x][y]["OSPF DIO metrics2"] = rtr_dict[y]["OSPF DIO metrics"]
			if 'list' in str(type(rtr_dict[y]["OSPF DIO metrics"])):
				write_cell(ws_summ,row_summ+1,6,list_to_str(rtr_dict[y]["OSPF DIO metrics"]))
			else:
				write_cell(ws_summ,row_summ+1,6,str(rtr_dict[y]["OSPF DIO metrics"]))
		
		print ("Checking for routes between x and y ", x, y)
		if "sir_parsed" in rtr_dict[x].keys() and "sir_parsed" in rtr_dict[y].keys():
			if "dict" in str(type (rtr_dict[x]["sir_parsed"])) and "dict" in str(type(rtr_dict[y]["sir_parsed"])):
				if "subnet_tree_routing_table" in rtr_dict[x]["sir_parsed"].keys() and "subnet_tree_routing_table" in rtr_dict[y]["sir_parsed"].keys():
					for network in rtr_dict[x]["sir_parsed"]['routing_table']:
						row = ws.max_row
						write_cell(ws, row+1, 1, network)
						#print ("Checking network 1", network)
						if network in rtr_dict[y]["sir_parsed"]["subnet_tree_routing_table"]:
							combo_op[x][y]["Matched"].append(network)
							next_hop1, raw_route_string1 = rtr_dict[x]["sir_parsed"]["subnet_tree_routing_table"][network]
							write_cell(ws, row+1, 2, str(raw_route_string1))
							next_hop2, raw_route_string2 = rtr_dict[y]["sir_parsed"]["subnet_tree_routing_table"][network]
							write_cell(ws, row+1, 3, str(raw_route_string2))
							#print ("route in both")
						else:
							next_hop1, raw_route_string1 = rtr_dict[x]["sir_parsed"]["subnet_tree_routing_table"][network]
							write_cell(ws, row+1, 2, str(raw_route_string1))
							combo_op[x][y]["Only in Rtr1"].append(network)
							#print ("route only in RTR1")
					for network in rtr_dict[y]["sir_parsed"]['routing_table']:
						#print ("Checking network 2", network)
						row = ws.max_row
						if network in rtr_dict[x]["sir_parsed"]["subnet_tree_routing_table"]:
							#print ("Route in both")
							continue
						else:
							write_cell(ws, row+1, 1, network)
							combo_op[x][y]["Only in Rtr2"].append(network)
							next_hop2, raw_route_string2 = rtr_dict[y]["sir_parsed"]["subnet_tree_routing_table"][network]
							write_cell(ws, row+1, 3, str(raw_route_string2))
							#print ("route only in RTR2")
			else:
				print ("x:",rtr_dict[x]["sir_parsed"],"\ny:",rtr_dict[y]["sir_parsed"])
		#print ("outside for loop")
		#print ("Matches Output in function : ", combo_op[x][y]["Matched"])
		#print ("Rtr1 Output in function : ", combo_op[x][y]["Only in Rtr1"])
		#print ("Rtr2 Output in function : ", combo_op[x][y]["Only in Rtr2"])
	wb.remove(wb['Sheet'])
	#if combo_op[x][y]["Matched"]:
		#freeze_panes (wb,"B1")
	wb.save(os.path.join(op_dir,region_site+".xlsx"))
	return nei_num, combo_op
			
def print_combo_op(combo_op,ws_routes,region,site,nei_num):
	row_routes = ws_routes.max_row
	site_count = 1
	if nei_num == 1:
		write_cell(ws_routes, row_routes+site_count, 1, region)
		write_cell(ws_routes, row_routes+site_count, 2, site)
		for i in combo_op.keys():
			write_cell(ws_routes, row_routes+site_count, 3, i)
		write_cell_style(ws_routes, row_routes+site_count, 4, "no neighbors",yellowFill)
		print ("Only one router at site", site)
		return 
	for rtr in combo_op.keys():
		o_metric1 = o_metric2 = ""
		for nei in combo_op[rtr].keys():
			print ("Printing to excel for rtr and nei ", rtr, nei)
			write_cell(ws_routes, row_routes+site_count, 1, region)
			write_cell(ws_routes, row_routes+site_count, 2, site)
			write_cell(ws_routes, row_routes+site_count, 3, rtr)
			write_cell(ws_routes, row_routes+site_count, 4, nei)
			if len(combo_op[rtr][nei]["Matched"]) > 0:
				write_cell(ws_routes, row_routes+site_count, 5, True)
			else:
				write_cell(ws_routes, row_routes+site_count, 5, False)
			if len(combo_op[rtr][nei]["Only in Rtr1"]) > 0:
				write_cell(ws_routes, row_routes+site_count, 6, True)
			else:
				write_cell(ws_routes, row_routes+site_count, 6, False)
			if len(combo_op[rtr][nei]["Only in Rtr2"]) > 0:
				write_cell(ws_routes, row_routes+site_count, 7, True)
			else:
				write_cell(ws_routes, row_routes+site_count, 7, False)
			if 'list' in str(type(combo_op[rtr][nei]["OSPF DIO metrics1"])):
				o_metric1 = list_to_str(combo_op[rtr][nei]["OSPF DIO metrics1"])
			else:
				o_metric1 = str(combo_op[rtr][nei]["OSPF DIO metrics1"])
			write_cell(ws_routes, row_routes+site_count, 8, o_metric1)
			if 'list' in str(type(combo_op[rtr][nei]["OSPF DIO metrics2"])):
				o_metric2 = list_to_str(combo_op[rtr][nei]["OSPF DIO metrics2"])
			else:
				o_metric2 = str(combo_op[rtr][nei]["OSPF DIO metrics2"])
			write_cell(ws_routes, row_routes+site_count, 9, o_metric2)
			o_metric1 = o_metric1.replace('nan','0')
			o_metric2 = o_metric2.replace('nan','0')
			write_cell(ws_routes, row_routes+site_count, 10, set([int(o_metric1), int(o_metric2)]).issubset(set([110, 120])))
			site_count+=1	
				
	return			
				
			

#Get child policy in parent policy
def get_plcy_and_child_defn(policy,pm_list):
	ch_pm = ""
	policy_defn = ""
	shaper = 0
	for i in pm_list:
		#print ("i in pol", i)
		if "policy-map "+policy in i:
			cpm = re.search("service-policy (\S+)",i)
			if cpm:
				ch_pm = cpm.group(1)
			policy_defn = i
			shp = re.search("shape average (\d+)",i)
			if shp:
				shaper = shp.group(1)
				policy_defn.replace(shp.group(0),"shape average")
				#print (policy_defn)
	for i in pm_list:
		if "policy-map "+ch_pm in i:
			policy_defn += i
	return policy_defn, shaper
	
	
def policy_check (policy_defn):
	policy_body = {"policy-map PM-DMVPN-OUT":0,"class CLASS-MAP-OUTPUT_COS_EF":0,"police cir percent 10":0,"set dscp tunnel ef":0,"priority level 1":0,"class CLASS-MAP-OUTPUT_COS_AF41":0,"set dscp tunnel af41":0,"bandwidth remaining percent 20":0,"random-detect dscp-based":0,"class CLASS-MAP-OUTPUT_COS_AF31":0,"set dscp tunnel af31":0,"bandwidth remaining percent 20":0,"random-detect dscp-based":0,"class CLASS-MAP-APPLICATION-MANAGEMENT":0,"set dscp tunnel default":0,"bandwidth remaining percent 10":0,"random-detect dscp-based":0,"class class-default":0,"random-detect":0,"bandwidth remaining percent 40":0,"policy-map PM-INTERNET-OUT":0,"class class-default":0,"shape average":0,"service-policy PM-DMVPN-OUT":0,"!":0}
	for i  in policy_defn:
		if i in policy_body:
			policy_body[i] +=1
	#print policy_body
	return
			

#convert column number to string
def colnum_string(n):
	string = ""
	while n > 0:
		n, remainder = divmod(n - 1, 26)
		string = chr(65 + remainder) + string
	return string

#convert list to a string separated with commas
#def list_to_str(list1):
#	list2 = list(set(list1))
#	str1 = str(list2).strip('[').strip(']')
#	str1.encode('ascii', 'ignore')
#	str1.replace("'","")
#	return str1
	
def list_to_str(list1):
	if len(list1) == 0:
		return ""
	str1 = ""
	for ele in list1:
		if ele == "Interface":
			continue
		if not str1:
			str1 = str(ele)
		else:
			str1 += ","+str(ele)
	return str1
	

#mask conversion
def mask(dot_mask):
	return sum([bin(int(x)).count('1') for x in str(dot_mask).split('.')])
	
	
def convert_netmask_to_prefix_length(mask_or_pref):
	if not mask_or_pref:
		return ""
	if re.match("^\/\d\d?$", mask_or_pref):
		return mask_or_pref
	if re.match("^\d\d?\d?\.\d\d?\d?\.\d\d?\d?\.\d\d?\d?$", mask_or_pref):
		return ("/"+ mask(mask_or_pref))
	return""

	
def network_stmt_convert_netmask_to_prefix_length(mask_or_pref):
	if not mask_or_pref:
		return ""
	if re.match("^\/\d\d?$", mask_or_pref):
		return mask_or_pref
	if re.match("^\d\d?\d?\.\d\d?\d?\.\d\d?\d?\.\d\d?\d?$", mask_or_pref):
		return ("/"+ str(32-mask(mask_or_pref)))
	return""

#writing with highlight
def write_cell(ws,row,col,value):
	ws.cell(row,col,value)
	if value == "WAN":
		ws[colnum_string(col) + str(row)].fill = yellowFill
	elif value == "" :
		ws[colnum_string(col) + str(row)].fill = yellowFill
	elif "0.0.0.0/0" in str(value):
		ws[colnum_string(col) + str(row)].fill = redFill
	elif "^r" in str(value):
		ws[colnum_string(col) + str(row)].fill = blueFill

def write_cell_style(ws,row,col,value,style):
	ws.cell(row,col,value)
	if style:
		ws[colnum_string(col) + str(row)].fill = style
	elif value == "WAN":
		ws[colnum_string(col) + str(row)].fill = yellowFill
	elif value == "" :
		ws[colnum_string(col) + str(row)].fill = yellowFill
	elif "0.0.0.0/0" in str(value):
		ws[colnum_string(col) + str(row)].fill = redFill
	elif "^r" in str(value):
		ws[colnum_string(col) + str(row)].fill = blueFill

def eigrp_topology_route(str_value,route):
	def_route_p = ""
	if str_value != "":
		rid = re.search("ID\(([\d\.]+)\)",str_value)
		adv_rid = re.findall("Originating router is ([\d\.]+)",str_value)
		if rid and len(adv_rid)>0:
			for adv_rtr in adv_rid:
				if rid.group(1) != adv_rtr:
					def_route_p = "Route to "+route+" in EIGRP topology table\n"
		if "Entry .* not in topology table" in str_value:
			def_route_p = "Route NOT in EIGRP topology table\n"
	else:
		def_route_p = "Route to "+route+" NOT in EIGRP topology table\n"
		
	return def_route_p

def ospf_database_route(str_value,route):
	def_route_p = ""
	if str_value != "":
		rid = re.search("ID \(([\d\.]+)\)",str_value)
		adv_rid = re.findall("Advertising Router: ([\d\.]+)",str_value)
		if rid and len(adv_rid)>0:
			for adv_rtr in adv_rid:
				if rid.group(1) != adv_rtr:
					def_route_p = "Route to "+route+" in OSPF database external table\n"
		if "LS age" not in str_value:
			def_route_p = "Route to "+route+" NOT in OSPF database external table\n"
	else:
		def_route_p = "Route to "+route+" NOT in OSPF database external table\n"
	
	return def_route_p

def ip_bgp_route(str_value,route): #Outliers - check if locally orignated 
	route_p = ""
	if "BGP not active" in str_value:
		route_p = "BGP not active\n"
	elif "Network not in table" in str_value:
		route_p = "Network "+route+" not in table\n"
	else:
		rt_p = ""
		for line in str_value.split("\n"):
			from_rt = re.search("BGP routing table entry for ([\d+\.\/]+)",line)
			from_nei = re.search("[\d+\.]+ from ([\d+\.]+)", line)
			origin = re.search("Origin (\w+)",line)
			if from_rt:
				rt_p += "Route to "+route+" from "+from_rt.group(1)+"\n"
			if from_nei:
				rt_p += "Route received from nei "+from_nei.group(1)+"\n"
				if from_nei.group(1)=="0.0.0.0":
					rt_p+="Route is Locally Originated"+"\n"
			if origin:
				rt_p += "Route origin "+origin.group(1)+"\n\n\n"
		if "172.16.0." in rt_p:
			route_p+= "Route to "+route+" advertised from DMVPN hub\n"+rt_p
		else:
			route_p+=rt_p
	
	return route_p
	
def sh_ip_route(str_value,route):
	route_p = ""
	if "Network not in table" in str_value or "Subnet not in table" in str_value:
		route_p = "Network "+route+" not in table\n"
	else:
		rt_p = ""
		for line in str_value.split("\n"):
			from_rt = re.search("Known via  \"(.*)\"",line)
			from_nei = re.search("[\d+\.]+, from ([\d+\.]+)", line)
			if from_rt:
				rt_p += "Route to "+route+" via "+from_rt.group(1)+"\n"
			if from_nei:
				rt_p += "Route received from nei "+from_nei.group(1)+"\n"
		if "172.16.0." in rt_p:
			route_p+= "Route to "+route+" advertised from DMVPN hub\n"+rt_p
		else:
			route_p+=rt_p
	
	return route_p

#Check which protocol is enabled on interface 
def nw_in_proto(bgp_lines, rip_lines, ospf_lines, eigrp_lines, route_lines, static_lines, static_route_default_lines):
	network_subtree = SubnetTree.SubnetTree()
	network_stmts = {}
	bgp_nw = re.findall("network ((\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+))",bgp_lines)
	for item in bgp_nw:
		print ("bgp item: ",item)
		subnet = (item[1] + network_stmt_convert_netmask_to_prefix_length(item[2]))
		print ("bgp subnet: ", subnet)
		network_subtree[subnet] = "bgp"
		network_stmts[subnet] = "bgp"
	ospf_nw = re.findall("network ((\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+))",ospf_lines)
	for item in ospf_nw:
		print ("ospf item: ",item)
		subnet = (item[1] + network_stmt_convert_netmask_to_prefix_length(item[2]))
		print ("ospf subnet: ", subnet)
		network_subtree[subnet] = "ospf"
		network_stmts[subnet] = "ospf"
	rip_nw = re.findall("network ((\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+))",rip_lines)
	for item in rip_nw:
		print ("rip item: ",item)
		subnet = (item[1] + network_stmt_convert_netmask_to_prefix_length(item[2]))
		print ("rip subnet: ", subnet)
		network_subtree[subnet] = "rip"
		network_stmts[subnet] = "ospf"
	eigrp_nw = re.findall("network ((\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+))",eigrp_lines)
	for item in eigrp_nw:
		print ("eigrp item: ",item)
		subnet = (item[1] + network_stmt_convert_netmask_to_prefix_length(item[2]))
		print ("eigrp subnet: ", subnet)
		network_subtree[subnet] = "eigrp"
		network_stmts[subnet] = "eigrp"
		
	print (network_stmts,network_subtree)

	return network_stmts,network_subtree
	
def check_ip_in_proto(network_to_chk, network_stmts, network_subtree):
	
	protocol = ""
	if network_to_chk in network_subtree:
		protocol = network_subtree[network_to_chk]
	
	print ("Protocol for interface :",protocol )
	
	return protocol
			
			
#Parse and get all details of R1 for store
def rtr_dict(hn,wb,op_dir,masterTracker, policyData, dict_op_hn, source, ws_summ , ws,ws_int,ws_rtng,ws_plcy,ws_cdp,ws_wan,ws_dhcp,ws_sdwan,wan_nei_ip,ws_lc, ws_sdwan_rtr, ws_sdwan_site):
	
	#print ("in function for ", hn, "dict :", dict_op_hn)
	
	cluster = dict_op_hn['Cluster']
	country = dict_op_hn['Country']
	wrk_loc = dict_op_hn['Work Location']
	wrk_loc_code = dict_op_hn['Work Location Code']
	
	site_name = dict_op_hn["CMSP - Site"]
	region = dict_op_hn['CMSP - Region']
	dev_ip = dict_op_hn['Device IP']
	dev_domain = dict_op_hn['Device Name & Domain']
	
	wan_neighbors = dict_op_hn["wan_neighbors"]
	wan_nei_ip = dict_op_hn["wan_nei_ip"]
	
	if "show run" in dict_op_hn.keys():
		print ("show run type in dict",type(dict_op_hn["show run"]), "source : == ",source)
		#print (" show run op: ", dict_op_hn["show run"])
		if 'str' in str(type(dict_op_hn["show run"])):
			sh_run = dict_op_hn["show run"].split("\n")
		elif 'list' in str(type(dict_op_hn["show run"])):
			sh_run = dict_op_hn["show run"]
		else:
			print ("show run type not recognised")
			return dict_op_hn
		print ("show run type ",type(sh_run))
		#print ("Show run ", sh_run)
	else:
		print ("Show run not found in dictionary ", dict_op_hn.keys())
		return dict_op_hn
		
	if "show cdp nei" in dict_op_hn.keys():
		if "CORAL" in source:
			cdp_file = dict_op_hn["show cdp nei"].split("\n")
		else:
			cdp_file = dict_op_hn["show cdp nei"]

	if "show cdp neighbors" in dict_op_hn.keys():
		if "CORAL" in source:
			cdp_file = dict_op_hn["show cdp neighbors"].split("\n")
		else:
			cdp_file = dict_op_hn["show cdp neighbors"]
	#Create sheets for output

	#Initialize variables
	dev_type = []
	intf_name = ""
	rtr_bgp = False
	lo0 = False
	mgmt_ip = ""
	tun_ip = ""
	
	row_summ = ws_summ.max_row
	row = ws.max_row
	row_int = ws_int.max_row
	row_rtng = ws_rtng.max_row
	row_plcy = ws_plcy.max_row
	row_cdp = ws_cdp.max_row
	row_wan = ws_wan.max_row
	row_dhcp = ws_dhcp.max_row
	row_sdwan = ws_sdwan.max_row
	row_lc = ws_lc.max_row
	row_sdwan_rtr = ws_sdwan_rtr.max_row
	row_sdwan_site = ws_sdwan_site.max_row
	
	#SDWAN RTR assessment
	if row_sdwan_rtr == 1:
	
		ws_sdwan_rtr.cell(1,1,"Device Details")
		#ws_sdwan_rtr.merge_cells('A1:E1')
		ws_sdwan_rtr.cell(2,1,"Country")
		ws_sdwan_rtr.cell(2,2,"Work Location")
		ws_sdwan_rtr.cell(2,3,"Comments")
		ws_sdwan_rtr.cell(2,4,"Hostname")
		
		ws_sdwan_rtr.cell(1,5,"Site SDWAN Readiness Status")
		ws_sdwan_rtr.cell(2,5,"SDWAN Migration Ready")
		ws_sdwan_rtr.cell(2,6,"HW Mandatory Ready")
		ws_sdwan_rtr.cell(2,7,"HW optional Ready")
		ws_sdwan_rtr.cell(2,8,"SW/CLI mandatory ready")
		ws_sdwan_rtr.cell(2,9,"SW/CLI optional Ready")
		ws_sdwan_rtr.cell(2,10,"HW remediation required (Mandatory)")
		ws_sdwan_rtr.cell(2,11,"HW remediation required (Optional)")
		ws_sdwan_rtr.cell(2,12,"SW Remediation Required (joint decision Cisco and UL) (Mandatory)")
		ws_sdwan_rtr.cell(2,13,"Mitigate during migration (SW or CLI based) (mandatory - timeline to be decided)")
		
		ws_sdwan_rtr.cell(1,14,"Circuit Checks")
		ws_sdwan_rtr.cell(2,14,"Circuit Type")
		ws_sdwan_rtr.cell(2,15,"Are circuits ready for migration to SDWAN as per SDD ?") 
		
		ws_sdwan_rtr.cell(1,16,"HW checks")
		#ws_sdwan_rtr.merge_cells('F1:J1')
		ws_sdwan_rtr.cell(2,16,"Chassis")
		ws_sdwan_rtr.cell(2,17,"Chassis support for SDWAN")
		ws_sdwan_rtr.cell(2,18,"LC(s)")
		ws_sdwan_rtr.cell(2,19,"LCs supported") ###
		ws_sdwan_rtr.cell(2,20,"LC(s) support for SDWAN")
		ws_sdwan_rtr.cell(2,21,"GE interfaces")
		ws_sdwan_rtr.cell(2,22,"GE interfaces more than 3 ?") ####
		
		ws_sdwan_rtr.cell(1,23,"HW recommended according to BW")
		#ws_sdwan_rtr.merge_cells('Y1:AA1')
		ws_sdwan_rtr.cell(2,23,"Max BW on device(MBPS)")
		ws_sdwan_rtr.cell(2,24,"Recommended Chassis for SDWAN by BW")	
		ws_sdwan_rtr.cell(2,25,"Recommended Chassis match existing HW")	
		
		ws_sdwan_rtr.cell(1,26,"Memory checks")
		#ws_sdwan_rtr.merge_cells('K1:P1')
		ws_sdwan_rtr.cell(2,26,"Bootflash total from show version (GB)")
		ws_sdwan_rtr.cell(2,27,"Bootflash total space ok for SDWAN")
		ws_sdwan_rtr.cell(2,28,"Bootflash free memory(GB)")
		ws_sdwan_rtr.cell(2,29,"Bootflash free aligned to current SDD for SDWAN")
		ws_sdwan_rtr.cell(2,30,"RAM (GB)")
		ws_sdwan_rtr.cell(2,31,"RAM aligned to current SDD for SDWAN")
		
		ws_sdwan_rtr.cell(1,32,"SW image and rommon version checks")
		#ws_sdwan_rtr.merge_cells('Q1:V1')
		ws_sdwan_rtr.cell(2,32,"Current SW version")
		ws_sdwan_rtr.cell(2,33,"SW image aligned with current SDD for SDWAN") # check how to do
		ws_sdwan_rtr.cell(2,34,"K9 image available ?")
		ws_sdwan_rtr.cell(2,35,"Rommon version")
		ws_sdwan_rtr.cell(2,36,"Rommon aligned to current SDD for SDWAN")
		ws_sdwan_rtr.cell(2,37,"Recommended Rommon version for HW for SDWAN 17.2")
		
		ws_sdwan_rtr.cell(1,38,"IGP checks")
		#ws_sdwan_rtr.merge_cells('AB1:AC1')	
		ws_sdwan_rtr.cell(2,38,"IGP")
		ws_sdwan_rtr.cell(2,39,"IGP supported for SDWAN ?")
		
		ws_sdwan_rtr.cell(1,40,"B2B with WAN neighbor")
		#ws_sdwan_rtr.merge_cells('AB1:AC1')	
		ws_sdwan_rtr.cell(2,40,"B2B link with each WAN neighbor(for more details check P2P sheet)?")
		ws_sdwan_rtr.cell(2,41,"CDP enabled on device?")
		
		ws_sdwan_rtr.cell(1,42,"HSRP checks")
		#ws_sdwan_rtr.merge_cells('W1:X1')
		ws_sdwan_rtr.cell(2,42,"No change required for HSRP for migration to SDWAN ?")
		ws_sdwan_rtr.cell(2,43,"HSRP interfaces")
		
		ws_sdwan_rtr.cell(1,44,"ZBFW checks")
		ws_sdwan_rtr.cell(2,44,"ZBFW zones")
		ws_sdwan_rtr.cell(2,45,"ZBFW zones as per SDD")
		
		ws_sdwan_rtr.cell(1,46,"Voice checks")
		#ws_sdwan_rtr.merge_cells('AE1:AG1')
		ws_sdwan_rtr.cell(2,46,"Is Voice CLI present ?")
		ws_sdwan_rtr.cell(2,47,"Are Voice cards present ?")
		ws_sdwan_rtr.cell(2,48,"Voice status on device ?")
		ws_sdwan_rtr.cell(2,49,"No voice present on device ?")
		
		ws_sdwan_rtr.cell(1,50,"LTE checks")
		ws_sdwan_rtr.cell(2,50,"LTE cards present")
		
		ws_sdwan_rtr.cell(1,51,"ATM checks")
		ws_sdwan_rtr.cell(2,51,"ATM cards present")
		
		ws_sdwan_rtr.cell(1,52,"Certificate information")
		#ws_sdwan_rtr.merge_cells('AJ1:AL1')	
		ws_sdwan_rtr.cell(2,52,"Certificate number")
		ws_sdwan_rtr.cell(2,53,"PID")
		ws_sdwan_rtr.cell(2,54,"Serial Number")
		ws_sdwan_rtr.cell(2,55,"Certificate and serial available ?")
		
		ws_sdwan_rtr.cell(1,56,"Interface Description")
		ws_sdwan_rtr.cell(2,56,"Interface description")
		ws_sdwan_rtr.cell(2,57,"Description correct as per standard ?")
		
		ws_sdwan_rtr.cell(1,58,"OSPF DIO metrics")
		ws_sdwan_rtr.cell(2,58,"OSPF DIO metrics")
		ws_sdwan_rtr.cell(2,59,"DIO metrics standard ?")
		
		ws_sdwan_rtr.cell( 2, 60, "Cluster")
		ws_sdwan_rtr.cell( 2, 61, 'Work Location Code')
		ws_sdwan_rtr.cell( 2, 62, 'CMSP - Site')
		ws_sdwan_rtr.cell( 2, 63, 'CMSP - Region')
		ws_sdwan_rtr.cell( 2, 64, 'Device IP')
		ws_sdwan_rtr.cell( 2, 65, 'Device Name & Domain')
		ws_sdwan_rtr.cell( 2, 66, 'Free interface for cross-connect available')


	
	row_sdwan_rtr = ws_sdwan_rtr.max_row
	
	rip = ospf = eigrp = bgp = static = False
	rout_parameters = ["passive-interface", "redistribute", "default-information", "network", "ip route", "router rip", "router ospf", "router bgp", "router eigrp", "router isis" ]
	
	sdwan_supported_hw = ["ASR1001-X", "ASR1002-X", "ASR1001-HX", "ASR1002-HX", "ISR4531/K9", "ISR4451-X/K9", "ISR4451/K9", "ISR4431/K9", "ISR4351/K9", "ISR4331/K9"]
	sdwan_isr4k_lc_support = ['NIM-1GE-CU-SFP','NIM-2GE-CU-SFP','NIM-1MFT-T1/E1','NIM-2MFT-T1/E1','NIM-4MFT-T1/E1','NIM-8MFT-T1/E1','NIM-ES2-4','NIM-ES2-8','NIM-LTEA-EA','NIM-LTEA-LA','NIM-VAB-A','NIM-VAB-M','SM-X-4X1G-1X10G','SM-X-6X1G','NIM-ES2-8-P','NIM-2FXO','NIM-4FXO','NIM-2FXSP','NIM-4FXSP','NIM-2FXS/4FXOP','SM-X-24FXS/4FXO','SM-X-16FXS/2FXO','SM-X-8FXS/12FXO','SM-X-72FXS','NIM-1T']
	sdwan_ver_17_2 = {"ASR1002-HX":"16.9(4)","ASR1001-HX":"16.9(4)","ASR1001-X":"16.9(4)","ASR1002-X":"16.7(1r)","ISR4321/K9":"16.12(2r)","ISR4331/K9":"16.12(2r)","ISR4351/K9":"16.12(2r)","ISR4431/K9":"16.12(2r)","ISR4451/K9":"16.12(2r)","ISR4461/K9":"16.12(2r)","ISR4451-X/K9":"16.12(1r)"}
	chassis_bw_map = {"ISR4331/K9":{"from":-1,"to":40}, "ISR4351/K9":{"from":40,"to":70}, "ISR4431/K9":{"from":70,"to":100}, "ISR4451/K9":{"from":100,"to":200}, "ISR4451-X/K9":{"from":100,"to":200}, "ASR1001-X":{"from":200,"to":1000}, "ASR1001-HX":{"from":1000,"to":10000}, "ASR1002-HX":{"from":1000,"to":10000}}
	voice_cards = ["WIC-2T","PVDM2-16","WIC-2T","HWIC-4B-S/T","PVDM4-32","HWIC-1T","HWIC-2CE1T1-PRI","WIC-1B-S/T-V3","WIC-2T=","HWIC-4T","HWIC-2T","VWIC3-1MFT-T1/E1","PVDM3-32","PVDM4-64"]
	lte_cards = ["EHWIC-4G-LTE-A","EHWIC-4G-LTE-AT"]
	atm_cards = ["EHWIC-4SHDSL-EA","NM-1A-OC3-POM"]
	
	
	sdwan_isr4k_lc_support_port_info = {'NIM-1GE-CU-SFP':1,'NIM-2GE-CU-SFP':2,'NIM-1MFT-T1/E1':1,'NIM-2MFT-T1/E1':2,'NIM-4MFT-T1/E1':4,'NIM-8MFT-T1/E1':8,'NIM-ES2-4':4,'NIM-ES2-8':8,'NIM-LTEA-EA':1,'NIM-LTEA-LA':1,'NIM-VAB-A':1,'NIM-VAB-M':1,'SM-X-4X1G-1X10G':4,'SM-X-6X1G':6,'NIM-ES2-8-P':8,'NIM-2FXO':2,'NIM-4FXO':4,'NIM-2FXSP':2,'NIM-4FXSP':4,'NIM-2FXS/4FXOP':2,'SM-X-24FXS/4FXO':24,'SM-X-16FXS/2FXO':16,'SM-X-8FXS/12FXO':8,'SM-X-72FXS':72,'NIM-1T':1}
	
	
	voice_params = ["voice-card"]
	voice = False
	voice_lines = ""
	
	nat_info = []
	bandwidth = []
	circuit_id = []
	wan_intf = []
	
	dhcp = False
	dhcp_lines = {}
	dhcp_pool = ""

	class_map_list = []
	policy_map_list = []
	acl_list = []
	cm = False
	pm = False
	acl = False
	cmlines = ""
	pmlines = ""
	acllines = ""
	bgp_lines = rip_lines = ospf_lines = eigrp_lines = route_lines = static_lines = static_route_default_lines = ""
	bgp_dio=ospf_dio=eigrp_dio=rip_dio = ""
	start_intf = False
	int_dict = {}
	lan_wan_dict = {}
	route_protocols = []
	zbfw = False
	zbfw_list = []
	track_lines = ip_sla_lines = ""
	icmp_echo_lines=""
	icmp_echo_count=0
	frequency_ipsla=""
	frequency_ipsla_count=0
	history_ipsla=""
	history_ipsla_count=0
	track = False
	ip_sla = False
	def_track_route=False
	tftp_lines = ftp_lines = telnet_lines = ""
	
	dev_max_bw = 0
	
	#Find variables in show run
	#sh_run.seek(0, 0)
	for line in sh_run:
		##print " Parsing show run"
		track_check = re.search("^track \d+ ",line)
		if track_check:
			##print "IP SLA"
			track = True
		if track:
			track_lines+=line
		ip_sla_check = re.search("^ip sla ",line)
		if ip_sla_check:
			ip_sla = True
		if ip_sla:
			#if line.lstrip() != line:
			if re.search("^\w+", line) and not re.search("^ip sla",line):
				ip_sla = False
			else:
				ip_sla_lines+=line
		if "icmp-echo" in line:
			icmp_echo_lines+=line
			icmp_echo_count+=1
		if "frequency" in line:
			frequency_ipsla+=line
			frequency_ipsla_count+=1
		if "history hours-of-statistics-kept" in line:
			history_ipsla+=line
			history_ipsla_count+=1
		if "ip route 0.0.0.0" in line and "track 100" in line:
			def_track_route=True
		if "zone-pair security" in line:
			zbfw = True
		zbfw_1 = re.search("zone security (\S+)", line)
		if zbfw_1:
			zbfw_list.append(zbfw_1.group(1))
		if "class-map match" in line:
			if cm:
				class_map_list.append(cmlines)
				cmlines = ""
			cm = True
		if cm:
			cmlines+=line
		if "policy-map" in line:
			if pm:
				policy_map_list.append(pmlines)
				pmlines = ""
			pm = True
		if pm:
			pmlines+=line
		if "access-list" in line:
			if acl:
				acl_list.append(acllines)
				acllines = ""
			acl = True
		if "permit" not in line and "forward-protocol" not in line:
			if "tftp" in line :
				tftp_lines = tftp_lines+line
			elif "ftp" in line:
				ftp_lines = ftp_lines+line
			if "telnet" in line or "^line" in line:
				telnet_lines = telnet_lines+line
		if acl:
			acllines+=line
		HN2 = re.search("hostname ([\w\-]+)", line)
		nat = re.search("ip nat pool (\S+) (\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+) netmask (\d+\.\d+\.\d+\.\d+)", line)
		if nat:
			nat_info.append(nat.groups())
			#nat_ip_mask_cidr.append("/"+ str(mask(nat_ip_mask)))
		if HN2:
			hn = HN2.group(1)	
			##print"getting hostname from device"	
		intf_line = re.search("^interface (\S+\s*[\/\.\d]+)",line)
		if intf_line:
			##print " Found intf: ",intf_line.group(1)
			intf_name = intf_line.group(1)
			start_intf = True
			int_lines = ""

		if start_intf:
			int_lines+=line+"\n"

				
		#ip_add = re.search("^ ip address (\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+)",line)
		#if ip_add and tun:
		#	if not tun_ip:
		#		tun_ip = ip_add.group(1)
		#	tun = False
			
		### Routing summary - redistribution ###
		
		if "vbond" in line:
			dev_type.append("VEdge")
		if "router rip" in line:
			route_protocols.append("RIP")
			rip = True
			rip_dio+=line
		if "router ospf" in line:
			route_protocols.append("OSPF")
			ospf = True
			ospf_dio+=line
		if "router eigrp" in line:
			route_protocols.append("EIGRP")
			eigrp = True
			eigrp_dio+=line
		if "router bgp" in line:
			route_protocols.append("BGP")
			bgp = True
			bgp_dio+=line
		if "ip route" in line:
			route_protocols.append("static")
			static = True
		
		for item in voice_params:
			if item in line:
				voice=True
		
		if voice == True:
			voice_lines+=line
			
		for item in rout_parameters:
			if item in line:
				if rip:
					rip_lines+=line
					if "default-information" in line:
						rip_dio+=line
				elif bgp:
					bgp_lines+=line
					if "default-information" in line:
						bgp_dio+=line
				elif ospf:
					ospf_lines+=line
					if "default-information" in line:
						ospf_dio+=line
				elif eigrp:
					eigrp_lines+=line
					if "default-information" in line:
						eigrp_dio+=line
				elif static:
					static_lines+=line
				else:
					route_lines+=line
		if re.match("ip route.*0\.0\.0\.0\s+0\.0\.0\.0",line):
			static_route_default_lines += line
		dhcp_chk = re.search("ip dhcp pool (\S+)", line)
		
		if dhcp_chk:
			dhcp = True
			dhcp_pool = dhcp_chk.group(1)
			dhcp_lines[dhcp_pool] = line
		if dhcp:
			dhcp_lines[dhcp_pool] += line

		# reset all variables at end of block
		end_of_block = re.search("^!",line)
		if end_of_block:
			tun = False
			lo0 = False
			ip_sla = False
			rip = ospf = eigrp = bgp = static = False
			if start_intf:
				int_lines+=line
				##print intf_name, int_lines
				int_dict[intf_name] = int_lines
				start_intf = False
				intf_name = ""
				int_lines = ""
			elif rtr_bgp:
				rtr_bgp = False
			elif cm:
				class_map_list.append(cmlines)
				cmlines = ""
				cm = False
			elif pm:		
				policy_map_list.append(pmlines)
				pmlines = ""
				pm = False	
			elif acl:
				acl_list.append(acllines)
				acllines = ""
				acl = False
			elif dhcp:
				dhcp =False
				dhcp_pool = ""
			elif voice:
				voice = False
			if track:
				track_lines+=line
				track = False
			if ip_sla:
				ip_sla+=line
				ip_sla = False
				
	#end of for loop for sh run
	
	
	network_stmts,network_subtree = nw_in_proto(bgp_lines, rip_lines, ospf_lines, eigrp_lines, route_lines, static_lines, static_route_default_lines)

	#write all variables to excel
	
	#Summary sheet
	if row_summ == 1:
		ws_summ.cell(row_summ,1,"Country")
		ws_summ.cell(row_summ,2,"Work Location")
		ws_summ.cell(row_summ,3,"Device IP")
		ws_summ.cell(row_summ,4,"Device Name and Domain")
		ws_summ.cell(row_summ,5,"Hostname")
		ws_summ.cell(row_summ,6,"Circuit Types seen")
		ws_summ.cell(row_summ,7,"DMVPN")
		ws_summ.cell(row_summ,8,"MPLS")
		ws_summ.cell(row_summ,9,"PIO")
		ws_summ.cell(row_summ,10,"IEN")
		ws_summ.cell(row_summ,11,"ISP router")
		ws_summ.cell(row_summ,12,"3G/4G cellular")
		ws_summ.cell(row_summ,13,"Agile router only")
		ws_summ.cell(row_summ,14,"Number of routers at site")
		ws_summ.cell(row_summ,15,"Production WAN resiliency configured")
		ws_summ.cell(row_summ,16,"MGMT WAN resiliency configured")
		ws_summ.cell(row_summ,17,"WAN routing protocol seen")
		ws_summ.cell(row_summ,18,"CORE routing protocol seen")
		ws_summ.cell(row_summ,19,"OSPF")
		ws_summ.cell(row_summ,20,"BGP")
		ws_summ.cell(row_summ,21,"EIGRP")
		ws_summ.cell(row_summ,22,"RIP")
		ws_summ.cell(row_summ,23,"Static")
		ws_summ.cell(row_summ,24,"Number of OSPF protocols")
		ws_summ.cell(row_summ,25,"Agile implemented")
		ws_summ.cell(row_summ,26,"HSRP on LAN")
		ws_summ.cell(row_summ,27,"Current HW platform")
		ws_summ.cell(row_summ,28,"Supported HW for DMVPN & SDWAN")
		ws_summ.cell(row_summ,29,"Current SW image")
		ws_summ.cell(row_summ,30,"Supported SW for SDWAN")
		ws_summ.cell(row_summ,31,"B2B between WAN")
		ws_summ.cell(row_summ,32,"WAN side config complete")
		ws_summ.cell(row_summ,33,"LAN side config complete")
		ws_summ.cell(row_summ,34,"MGMT config complete")
		ws_summ.cell(row_summ,35,"Firewall & Security config complete")
		ws_summ.cell(row_summ,36,"Base config complete")
		ws_summ.cell(row_summ,37,"BW command under interface")
		ws_summ.cell(row_summ,38,"BW ordered")
		ws_summ.cell(row_summ,39,"BW on the interface description")
		ws_summ.cell(row_summ,40,"ZBFW implemented ?")
		ws_summ.cell(row_summ,41,"ZBFW zones")
		
		
	#RTR summary
	if row == 1:
		ws.cell(row,1,"Country")
		ws.cell(row,2,"Work Location")
		ws.cell(row,3,"Hostname")
		ws.cell(row,4,"IP seen from device")
		ws.cell(row,5,"Device Type")
		ws.cell(row,6,"Routing in use")
		ws.cell(row,7,"Existing Tunnel IP")
		ws.cell(row,8,"Ordered net mask")
		ws.cell(row,9,"Existing net mask on WAN interfaces")
		ws.cell(row,10,"NAT Pool")
		ws.cell(row,11,"NAT IP start range")
		ws.cell(row,12,"NAT IP end range")
		ws.cell(row,13,"NAT Mask")
		ws.cell(row,14,"NAT Mask in CIDR")
		ws.cell(row,15,"ZBFW")
		ws.cell(row,16,"ZBFW list")
		ws.cell(row,17,"Source of analysis")
		ws.cell(row,18,"IP SLA")
		ws.cell(row,19,"Bandwidth of physcial interfaces")
		ws.cell(row,20,"Circuit ID seen")
		ws.cell(row,21,"WAN interface")
		ws.cell(row,22,"Device Type")
		ws.cell(row,23,"Image")
		ws.cell(row,24,"HSRP on LAN")
		ws.cell(row,25,"CDP neighbors")
		ws.cell(row,26,"WAN neighbor in CDP")
		ws.cell(row,27,"DHCP pool name(s)")
		ws.cell(row,28,"Maximum BW on device(MBPS)")
		ws.cell(row,29,"Ordered BW (MBPS)")
		ws.cell(row,30,"New Circuit Ref")
		ws.cell(row,31,"Device serial from Sh license udi")
		ws.cell(row,32,"Device UDI from Sh license udi")
		ws.cell(row,33,"CDP enabled")
	
	
	#Initializing the interface sheet
	if row_int == 1:
		ws_int.cell(row_int,1,"Country")
		ws_int.cell(row_int,2,"Work Location")
		ws_int.cell(row_int,3,"Hostname")
		ws_int.cell(row_int,4,"Interface")
		ws_int.cell(row_int,5,"Bandwidth")
		ws_int.cell(row_int,6,"IP address")
		ws_int.cell(row_int,7,"Network Mask")
		ws_int.cell(row_int,8,"Network Mask in CIDR")
		ws_int.cell(row_int,9,"Interface type")
		ws_int.cell(row_int,10,"Circuit Type")
		ws_int.cell(row_int,11,"Circuit ID")
		ws_int.cell(row_int,12,"Encapsulation")
		ws_int.cell(row_int,13,"Port in use")
		ws_int.cell(row_int,14,"NAT")
		ws_int.cell(row_int,15,"Standby config")
		ws_int.cell(row_int,16,"Routing protocol(s)")
		ws_int.cell(row_int,17,"Description")
		ws_int.cell(row_int,18,"Contents")
		##print " Int_row\n"
		
	if hn in masterTracker['Router Hostname'].unique():
		T_hn = masterTracker.loc[masterTracker['Router Hostname'] == hn]
		#["Unilever Circuit Reference","Cluster","Country","Site Name","Supplier (short)","Forecast/ Actual Migration Go Live Date","Circuit Port (Mbps)","RFP Reference","Supplier Prefix","IP Address Requirement"]]
		IP_add_req = str(T_hn['IP Address Ordered'].unique()[0])
		bw_req = str(T_hn['Ordered Bandwidth (Mbps)'].unique()[0])
		new_ckt_ref = str(T_hn['New Circuit Reference'].unique()[0])
		write_cell(ws, row+1,8,IP_add_req)
		write_cell(ws, row+1,29, bw_req)
		write_cell(ws, row+1, 30, new_ckt_ref)
	elif hn in masterTracker['New Router Hostname'].unique():
		T_hn = masterTracker.loc[masterTracker['New Router Hostname'] == hn]
		#["Unilever Circuit Reference","Cluster","Country","Site Name","Supplier (short)","Forecast/ Actual Migration Go Live Date","Circuit Port (Mbps)","RFP Reference","Supplier Prefix","IP Address Requirement"]]
		IP_add_req = str(T_hn['IP Address Ordered'].unique()[0])
		bw_req = str(T_hn['Ordered Bandwidth (Mbps)'].unique()[0])
		new_ckt_ref = str(T_hn['New Circuit Reference'].unique()[0])
		write_cell(ws, row+1,8,IP_add_req)
		write_cell(ws, row+1,29, bw_req)
		write_cell(ws, row+1, 30, new_ckt_ref)
	
	
	#Initialise DHCP sheet
	if row_dhcp == 1:
		ws_dhcp.cell(row_dhcp,1,"Country")
		ws_dhcp.cell(row_dhcp,2,"Work Location")
		ws_dhcp.cell(row_dhcp,3,"Hostname")
		ws_dhcp.cell(row_dhcp,4,"Pool name")
		ws_dhcp.cell(row_dhcp,5,"VRF")
		ws_dhcp.cell(row_dhcp,6,"dhcp_lines")	

	pool_count = 1
	for key, value in dhcp_lines.items():
		write_cell(ws_dhcp, row_dhcp+pool_count, 1, country)
		write_cell(ws_dhcp, row_dhcp+pool_count, 2, wrk_loc)
		write_cell(ws_dhcp, row_dhcp+pool_count, 3, hn)
		write_cell(ws_dhcp, row_dhcp+pool_count, 4, key)
		write_cell(ws_dhcp, row_dhcp+pool_count, 6, value)
		vrf_pool = re.search("vrf (\S+)",value)
		if vrf_pool:
			write_cell(ws_dhcp, row_dhcp+pool_count, 5, vrf_pool.group(1))
		pool_count+=1
		
	
	#Initialise the CDP sheet
	if row_cdp == 1:
		ws_cdp.cell(row_cdp,1,"Country")
		ws_cdp.cell(row_cdp,2,"Work Location")
		ws_cdp.cell(row_cdp,3,"Hostname")
		ws_cdp.cell(row_cdp,4,"WAN Neighbor")
		ws_cdp.cell(row_cdp,5,"CDP neighbor")
		ws_cdp.cell(row_cdp,6,"CDP enabled")
	
	
	write_cell(ws, row+1,3,hn)
	write_cell(ws, row+1,15,zbfw)
	write_cell(ws, row+1,16,list_to_str(zbfw_list))
	write_cell(ws, row+1,17,source)
	write_cell(ws, row+1,18, track_lines +"\n"+ip_sla_lines +"\n"+ static_route_default_lines)
	write_cell(ws, row+1,1, country)
	write_cell(ws, row+1,2, wrk_loc)
	if dhcp_lines:
		write_cell(ws, row+1,27, list_to_str(dhcp_lines.keys()))

	
	write_cell(ws_summ, row_summ+1,1, country)
	write_cell(ws_summ, row_summ+1,2, wrk_loc)
	write_cell(ws_summ, row_summ+1,3, dev_ip)
	write_cell(ws_summ, row_summ+1,4, dev_domain)
	write_cell(ws_summ, row_summ+1,5, hn)
	write_cell(ws_summ, row_summ+1,40,zbfw)
	write_cell(ws_summ, row_summ+1,41,list_to_str(zbfw_list))
	
	nei_presence = ""
	nei_count = 0
	cdp_enabled = False
	
	if wan_neighbors and cdp_file:
		
		for nei in wan_neighbors:
			#print (nei,"\n",cdp_file)
			nei_count=nei_count + 1
			write_cell(ws_cdp, row_cdp+nei_count, 1, country)
			write_cell(ws_cdp, row_cdp+nei_count, 2, wrk_loc)
			write_cell(ws_cdp, row_cdp+nei_count, 3, hn)
			write_cell(ws_cdp, row_cdp+nei_count, 4, nei)
			if nei in str(cdp_file):
				write_cell(ws_cdp, row_cdp+nei_count, 5, True)
				if not nei_presence:
					nei_presence = "True"
				else:
					nei_presence+=", True"
			else:
				write_cell(ws_cdp, row_cdp+nei_count, 5, False)
				if not nei_presence:
					nei_presence = "False"
				else:
					nei_presence+=", False"
			if "CDP is not enabled" in str(cdp_file) or str(cdp_file) == "":
				write_cell(ws_cdp, row_cdp+nei_count, 6, False)
				cdp_enabled = False
			else:
				write_cell(ws_cdp, row_cdp+nei_count, 6, True)
				cdp_enabled = True
	write_cell(ws, row+1,25, list_to_str(wan_neighbors) )
	write_cell(ws, row+1,26, str(nei_presence))
	write_cell(ws, row+1,33, cdp_enabled)
	
	if "False" in nei_presence:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,40,False )
	else:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,40,True )
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,41,cdp_enabled)
	
	if cdp_enabled:
		write_cell(ws_summ,row_summ+1,31,str(nei_presence))
	else:
		write_cell(ws_summ,row_summ+1,31,"CDP not enabled")
	
	elem_count = 1
	for element in nat_info:				
		write_cell(ws, row+elem_count,10, element[0])
		write_cell(ws, row+elem_count,11, element[1])
		write_cell(ws, row+elem_count,12, element[2])
		write_cell(ws, row+elem_count,13, element[3])
		write_cell(ws, row+elem_count,14, str(mask(element[3])))
		elem_count+=1
	net_mask = ""
	
	#print (int_dict.items())
	standby_in_rtr = False
	#iterate over each interface
	wan_interface_description = []
	circuit_det = {}
	free_int = []
	shut_int = {}
	for key, int_lines in int_dict.items():
		print ("Parsing interface", key)
		shut_int[key] = False
		subintf = re.search("([\w\d\/\/]+)\.(\d+)",key)
		if subintf:
			print ("Main intf : ", subintf.group(1))
			if subintf.group(1) in circuit_det.keys():
				c_type = circuit_det[subintf.group(1)]
				print ("Sub intf = ", key," c_type = ", c_type)
			if subintf.group(1) in shut_int.keys():
				shut_int[key] = shut_int[subintf.group(1)]
		
		#print ("Parsing value type ", str(type(int_lines)))
		#print (key)
		if "list" in str(type(int_lines)):
			list_value = int_lines
			value = list_to_str(int_lines)
			#print ("Type of content ", str(type(value)))
		else:
			list_value = str(int_lines).split("\n")
			value = int_lines
		#print ("Type of list content ", str(type(list_value)))
		#print ("Type of content ", str(type(value)))
		ip_add_intf = ""
		c_type = net_m = lan_wan = ip_nat = ""
		c_t = ""
		int_protocol = []
		tun_ip = []
		row_int = ws_int.max_row
		standby_lines = ""
		write_cell(ws_int, row_int+1, 1, country)
		write_cell(ws_int, row_int+1, 2, wrk_loc)
		write_cell(ws_int, row_int+1, 3, hn)
		write_cell(ws_int, row_int+1, 4, key)
		write_cell(ws_int, row_int+1, 18, value)
		if "Cellular" in key:
			write_cell(ws_summ, row_summ+1,12,"Yes")
		
		#protocol check:
		if "rip advertise" in value:
			int_protocol.append("rip")
		if "ip ospf" in value:
			int_protocol.append("ospf")

		if "unnel65534" in key: #for DMVPN tunnel
			c_type = "DMVPN"
			lan_wan = "WAN"
			#print ("DMVPN , WAN")
			
		if "ip nat inside" in value: #nat inside interface
			ip_nat = "inside"
			lan_wan = "LAN"
			#print ("NAT inside LAN")
		elif "ip nat outside" in value: #nat outside typically WAN interface
			ip_nat = "outside"
			lan_wan = "WAN"
			#print ("NAT outside WAN")

		if "ip vrf forwarding DMVPN-TRANSPORT-INTERNET" in value and "ip nat inside" not in value:
			c_type = "DMVPN"
			lan_wan = "WAN"
			#print ("DMVPN, WAN")
		

		#if "ip policy route-map " in value:
		#	lan_wan = lan_wan + " X-connect"

		if ("POLICY-MAP-MPLS_COS_SHAPE" in value or "POLICY-MAP-MPLS_COS_Shape" in value or "description.*MPLS.*" in value) and ("SDWAN" not in value and "underlay" not in value):
			if not c_type:
				c_type =  "MPLS"
				#print ("MPLS")
			if not lan_wan:
				lan_wan = "WAN"	
				#print("MPLS WAN")		
		if "x-over" in value or "X-over" in value:
			if not lan_wan:
				lan_wan = "X-connect"
				#print ("Xconnect")
		if "DMVPN" in  c_type and "LAN" in lan_wan: 
			lan_wan = "Agile"
			#print ("Agile")

		if lan_wan != "WAN":
			if "bandwidth " in value: #removed --> if not "unnel" in key
				lan_wan = "WAN"
				#print ("bandwidth WAN")
		#Check if the interface is in use

		if "GigabitEthernet" in key and "shutdown" in value:
			free_int+=key
		if "shutdown" in value:
			shut_int[key] = True
		elif "shutdown" in value and "no shutdown" not in value and "GigabitEthernet0" !=  key:
			write_cell(ws_int, row_int+1, 13, "No")
		else:
			write_cell(ws_int, row_int+1, 13, "Yes")
		if "shutdown" in value or shut_int[key]:
			write_cell(ws_int, row_int+1, 13, "Shutdown")
		if "no ip address" and ("shutdown" in value or shut_int[key]) and "GigabitEthernet0" !=  key:
			write_cell(ws_int, row_int+1, 13, "No")
		
		#For loop where individual lines are needed
		for int_line_ in list_value: 
			#find the bandwidth
			#print ("line: ", int_line_)
			bw = re.search("\s*bandwidth (\d+)",int_line_) 
			
			if bw:
				#print ("Found bandwidth", bw.group(1)) 
				write_cell(ws_int, row_int+1, 5, bw.group(1))
				if "lan" not in key and "BVI" not in key and "ailer" not in key:
					bw_calc = int(bw.group(1))/1000
					if bw_calc>dev_max_bw:
						dev_max_bw = bw_calc
					bw_str = ""
					if bw_calc <1:
						bw_str = str(bw.group(1)) + "Kbps"
					else:
						bw_str = str(bw_calc)+"MB"
					bandwidth.append(bw_str)
					#lan_wan = "WAN"
			
			#find the ip and mask
			ip_add_mask = re.search("ip address (\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+)",int_line_)
			if ip_add_mask:
				#print ("found ip address and mask")
				net_m = ip_add_mask.group(2)
				ip_add_intf = ip_add_mask.group(1)+"/"+str(mask(net_m))
				write_cell(ws_int, row_int+1, 6, ip_add_mask.group(1))
				write_cell(ws_int, row_int+1, 7, net_m)
				write_cell(ws_int, row_int+1, 8, str(mask(net_m)))
				
				int_protocol.append(check_ip_in_proto(ip_add_intf+"/"+str(mask(net_m)), network_stmts, network_subtree))
				
				if "unnel" in key: #assuming no internal tunnels
					tun_ip.append(ip_add_mask.group(1))
					if not c_type:
						c_type = "IEN"
						#print ("IEN")
					if not lan_wan:
						lan_wan = "WAN"
						#print ("IEN WAN")
			
			#find the encap
			encap = re.search("encapsulation (dot1Q \d+)",int_line_)
			if encap:
				write_cell(ws_int, row_int+1, 12, encap.group(1)) #dot1q encap vlan
			encap_fr = re.search("encapsulation (frame-relay \w+)", int_line_)
			if encap_fr:
				write_cell(ws_int, row_int+1, 12, encap_fr.group(1)) #FR encap
			encap_fr_dlci = re.search("(frame-relay interface-dlci \d+)",int_line_)
			if encap_fr_dlci:
				write_cell(ws_int, row_int+1, 12, encap_fr_dlci.group(1)) #FR encap
			if "standby " in int_line_: #check for HSRP config
				standby_lines = standby_lines+int_line_+"\n"
			c_t = ""
			#description parsing
			desc = re.search(" description (.*)", int_line_)
			if desc:
				#print ("Description found -->", desc.group(1))
				write_cell(ws_int, row_int+1, 17, desc.group(1))
				cid = re.search("ID\s*[\:\-\s]+\s*(\S+)",desc.group(1))
				if cid:
					##print " Found circuit id\n"
					write_cell(ws_int, row_int+1, 11, cid.group(1))
					circuit_id.append(cid.group(1))
				mpls = re.search("description .*(BT MPLS|BT GMPLS|MPLS).*", int_line_)
				if mpls and "no ip address" not in value and "LAN" not in desc.group(1) and "VLAN" not in desc.group(1) and "Agile" not in desc.group(1) and "DMVPN" not in desc.group(1) and ("SDWAN" not in desc.group(1) and "underlay" not in desc.group(1)) and "unnel" not in key:
					c_t = mpls.group(1)
					if not lan_wan and "unnel" not in key:
						lan_wan = "WAN"
						#print ("description MPLS")
				#elif "BT" in desc.group(1):
				#	c_type = "Check the interface"
				elif "LAN" in desc.group(1):
					lan_wan = "LAN"
				elif "DMVPN" in desc.group(1):
					if not lan_wan:
						c_type = "DMVPN"
				elif "BT Managemen" in desc.group(1) or "oopback0" in key or "anagemen" in desc.group(1) or "mgmt|MGMT|Mgmt" in desc.group(1):
					#if not c_type: #removed 21-sep AG
					c_type = "MGMT"
					lan_wan = "MGMT"
					#if not lan_wan:
					#	lan_wan = "MGMT"
				elif "oopback" in key:
					lan_wan = "LAN"
				elif "Agile" in desc.group(1) or "AGILE" in desc.group(1) or "agile" in desc.group(1):
					c_type = "Agile"
					if not lan_wan:
						lan_wan = "LAN"
				if "IP for NAT" in desc.group(1):
					c_type = "Public IP for NAT"
					lan_wan = "Public IP"
				if c_t:
					if c_type and c_type!=c_t:
						c_type += ","+c_t
					else:
						c_type = c_t
		if (c_type == "") and ("WAN" in lan_wan) and ("oopback" not in key) and ("ip nat inside" not in value) and ("no ip address" not in value) and ("shutdown" not in value):
			c_type = "PIO"
		elif not lan_wan:
			lan_wan = "LAN"
		print ("intf = ", key," c_type = ", c_type)
		ws_int.cell(row_int+1,14, ip_nat)
		ws_int.cell(row_int+1,10, c_type)
		circuit_det[key] = c_type
		ws_int.cell(row_int+1,16, list_to_str(int_protocol))
		ws_int.cell(row_int+1,15, standby_lines)
		write_cell(ws_int, row_int+1, 9, lan_wan)
		lan_wan_dict[key] = lan_wan
		if ip_add_intf:
			if ip_add_intf == dev_ip: #### 24-Jul AG: changed to dev_ip check
				mgmt_ip = ip_add_intf
				mgmt_intf = key
		if standby_lines:
			standby_in_rtr = True
		if "WAN" in lan_wan and "shutdown" not in value:
			dev_type.append(c_type)
			wan_intf.append(key)
			if desc:
				wan_interface_description.append(desc.group(1))
			#net_mask = mask(net_m)
			if net_mask and net_m:
				net_mask = net_mask + ",/" + str(mask(net_m))
				##print net_mask
			elif net_m:
				net_mask = "/"+str(mask(net_m))
				##print net_mask
	write_cell(ws, row+1,5,list_to_str(dev_type))
	write_cell(ws_summ, row_summ+1,6,list_to_str(list(set(dev_type))))
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,14,list_to_str(list(set(dev_type))))
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,66,list_to_str(free_int))
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,56,list_to_str(wan_interface_description))
	if "|" in list_to_str(wan_interface_description) and "^#" in list_to_str(wan_interface_description):
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,57,True)
	else:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,57,False)
	if "PIO" in dev_type or "IEN" in dev_type or "MPLS" in dev_type or not dev_type:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,15,False)
	else:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,15,True)
	gige_count =  sum('GigabitEthernet0/' in key for key in int_dict.keys())
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,21,str(gige_count))
	if gige_count>=3:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,22,True)
	else:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1,22,False)
	dmvpn_count = 0
	if dev_type.count("DMVPN") > 0:
		dmvpn_count = dev_type.count("DMVPN") -1
	write_cell(ws_summ, row_summ+1,7,str(dmvpn_count)) #subtracting the instance for tunnel65534
	write_cell(ws_summ, row_summ+1,8,str(dev_type.count("MPLS")))
	write_cell(ws_summ, row_summ+1,9,str(dev_type.count("PIO")))
	write_cell(ws_summ, row_summ+1,10,str(dev_type.count("IEN")))
	if "isp|ISP" in hn:
		write_cell(ws_summ, row_summ+1,11,"Yes")
	
	write_cell(ws, row+1,7,list_to_str(tun_ip))
	
	#f lan_wan == "WAN":
	#	dev_type.append(c_type)
	#	wan_intf.append(key)
	#	##print net_m
	#	if net_mask and net_m:
	#		net_mask = net_mask + ",/" + str(mask(net_m))
	#		##print net_mask
	#	elif net_m:
	#		net_mask = "/"+str(mask(net_m))
	#		##print net_mask
	write_cell(ws, row+1,9,net_mask)
	write_cell(ws, row+1,24,str(standby_in_rtr))
	
	write_cell(ws, row+1,4,mgmt_ip)	
	write_cell(ws, row+1,28, str(dev_max_bw))
	write_cell(ws_sdwan, row_sdwan+1, 19, str(dev_max_bw))
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 23, str(dev_max_bw)) #sdwan
	#print ("Dev max bw = ", str(dev_max_bw))
	#print(type(dev_max_bw))
	
	#Initialising the routing summary sheet
	if row_rtng == 1:
		ws_rtng.cell(row_rtng,1,"Country")
		ws_rtng.cell(row_rtng,2,"Work Location")
		ws_rtng.cell(row_rtng,3,"Hostname")
		ws_rtng.cell(row_rtng,4,"Routing Protocols in use")
		ws_rtng.cell(row_rtng,5,"BGP")
		ws_rtng.cell(row_rtng,6,"OSPF")
		ws_rtng.cell(row_rtng,7,"Static")
		ws_rtng.cell(row_rtng,8,"RIP(if exists)")
		ws_rtng.cell(row_rtng,9,"EIGRP(if exists)")
		ws_rtng.cell(row_rtng,10,"Other routing(if exists)")
		ws_rtng.cell(row_rtng,11,"Count of OSPF processes")
		ws_rtng.cell(row_rtng,12,"IP SLA CLI")
		ws_rtng.cell(row_rtng,13,"OSPF Default information originate")
		ws_rtng.cell(row_rtng,14,"BGP Default information originate")
		ws_rtng.cell(row_rtng,15,"RIP Default information originate")
		ws_rtng.cell(row_rtng,16,"EIGRP Default information originate")
		ws_rtng.cell(row_rtng,17,"OSPF Default information metric")
				
	write_cell(ws_rtng,row_rtng+1,1,country)
	write_cell(ws_rtng,row_rtng+1,2,wrk_loc)
	write_cell(ws_rtng,row_rtng+1,3,hn)
	write_cell(ws_rtng,row_rtng+1,4,list_to_str(np.unique(np.array(route_protocols))))
	write_cell(ws_rtng,row_rtng+1,5,bgp_lines)
	write_cell(ws_rtng,row_rtng+1,6,ospf_lines)
	write_cell(ws_rtng,row_rtng+1,7,static_lines)
	write_cell(ws_rtng,row_rtng+1,8,rip_lines)
	write_cell(ws_rtng,row_rtng+1,9,eigrp_lines)
	write_cell(ws_rtng,row_rtng+1,10,route_lines)
	write_cell(ws_rtng,row_rtng+1,13,ospf_dio)
	write_cell(ws_rtng,row_rtng+1,14,bgp_dio)
	write_cell(ws_rtng,row_rtng+1,15,rip_dio)
	write_cell(ws_rtng,row_rtng+1,16,eigrp_dio)
	
	ospf_dio_metrics = re.findall("default-information originate metric (\d+) metric-type 1",ospf_dio)
	if len(ospf_dio_metrics) > 0:
		write_cell(ws_rtng,row_rtng+1,17,list_to_str(ospf_dio_metrics))
		dict_op_hn["OSPF DIO metrics"] = []
		dict_op_hn["OSPF DIO metrics"] = ospf_dio_metrics
	
	#bgp_dio_metrics = re.findall("default-information originate metric (\d+) metric-type 1",bgp_dio)
	#rip_dio_metrics = re.findall("default-information originate metric (\d+) metric-type 1",rip_dio)
	#eigrp_dio_metrics = re.findall("default-information originate metric (\d+) metric-type 1",eigrp_dio)
	 
	#### 24/Jul AG: Identify WAN and LAN protocols
	
	ospf_count = ospf_lines.count("router ospf")
	write_cell(ws_rtng,row_rtng+1,11,str(ospf_count))
	write_cell(ws_rtng, row_rtng+1,12, track_lines +"\n"+ip_sla_lines+"\n"+static_route_default_lines)
	write_cell(ws_wan,row_wan+1,6,track_lines +"\n"+ip_sla_lines+"\n"+static_route_default_lines)
	
	write_cell(ws,row+1,6,list_to_str(np.unique(np.array(route_protocols))))
	
		
	#Initializing the policy sheet
	if row_plcy == 1:		
		ws_plcy.cell(row_plcy,1,"Country")
		ws_plcy.cell(row_plcy,2,"Work Location")
		ws_plcy.cell(row_plcy,3,"Hostname")
		ws_plcy.cell(row_plcy,4,"class-maps")
		ws_plcy.cell(row_plcy,5,"policy-maps")
		ws_plcy.cell(row_plcy,6,"ACLs")
		ws_plcy.cell(row_plcy,7,"WAN output policy")
		ws_plcy.cell(row_plcy,8,"Does it match recommended value of PM-INTERNET-OUT?")
		ws_plcy.cell(row_plcy,9,"Shaper BW value")
		ws_plcy.cell(row_plcy,10,"Telnet lines")
		ws_plcy.cell(row_plcy,11,"TFTP lines")
		ws_plcy.cell(row_plcy,12,"FTP lines")
	
	ab=cd=shaper = ""
	ab,shaper = get_plcy_and_child_defn("PM-INTERNET-OUT",policy_map_list)
	cd = policy_check(ab)
	pol_chk = re.search("shape average (\d+)",ab)
	if pol_chk:
		ab.replace(pol_chk.group(0),"shape average")
		#print (ab)
	write_cell(ws_plcy, row_plcy+1, 1, country)
	write_cell(ws_plcy, row_plcy+1, 2, wrk_loc)
	write_cell(ws_plcy, row_plcy+1, 3, hn)
	write_cell(ws_plcy, row_plcy+1, 4, list_to_str(class_map_list))
	write_cell(ws_plcy, row_plcy+1, 5, list_to_str(policy_map_list))
	write_cell(ws_plcy, row_plcy+1, 6, list_to_str(acl_list))
	write_cell(ws_plcy, row_plcy+1, 7, ab)
	write_cell(ws_plcy, row_plcy+1, 8, str(policyData == ab))
	write_cell(ws_plcy, row_plcy+1, 9, str(shaper))
	write_cell(ws_plcy, row_plcy+1, 10, telnet_lines)
	write_cell(ws_plcy, row_plcy+1, 11, tftp_lines)
	write_cell(ws_plcy, row_plcy+1, 12, ftp_lines)
	write_cell(ws,row+1,19,list_to_str(bandwidth))
	write_cell(ws,row+1,20,list_to_str(circuit_id))
	
	#get device type and Image from metadata of NP
	device_type = ""
	sys_desc = ""
	if "NP" in source and "metadata" in dict_op_hn.keys():
		##print " Found metadata file"
		dev_t = re.search("\<device_type\>(\S+)\<\/device_type\>",str(dict_op_hn["metadata"]))
		if dev_t:
			##print " Found Device type"
			device_type = dev_t.group(1)
			write_cell(ws,row+1,22,device_type)
		#sys_d = re.search("\<sys_description\>.+Version (.+),.+</sys_description>", line)
		sys_d = re.search("\<os_version\>(.+)\</os_version\>",str(dict_op_hn["metadata"]))
		if sys_d:
			##print " Found Image"
			sys_desc = sys_d.group(1)
			write_cell(ws,row+1,23,sys_desc)
	if "show version" in dict_op_hn.keys():
		version = re.findall("Version [\d\w\.\S\(\)]+\,?\s+",dict_op_hn["show version"])
		if version and 'list' in str(type(version)): 
			write_cell(ws,row+1,23,list_to_str(version[0:2]))
			write_cell(ws_sdwan_rtr,row_sdwan_rtr+1,32,list_to_str(version[0:2]))
			version_check = 0
			version_check = sum("16.9.5" in s for s in version[0:2])
			if version_check>0:
				write_cell(ws_sdwan_rtr,row_sdwan_rtr+1,33,True)
			else:
				write_cell(ws_sdwan_rtr,row_sdwan_rtr+1,33,False)
		if "K9" in dict_op_hn["show version"]:
			write_cell(ws_sdwan_rtr,row_sdwan_rtr+1,34,True)
		else:
			write_cell(ws_sdwan_rtr,row_sdwan_rtr+1,34,False)
	write_cell(ws,row+1,21,list_to_str(wan_intf))
	
	#SDWAN assessment
	
	if row_sdwan == 1:
		ws_sdwan.cell(row_sdwan,1,"Country")
		ws_sdwan.cell(row_sdwan,2,"Work Location")
		ws_sdwan.cell(row_sdwan,3,"Hostname")
		ws_sdwan.cell(row_sdwan,4,"Bootflash total memory(GB)")
		ws_sdwan.cell(row_sdwan,5,"Bootflash free memory(GB)")
		ws_sdwan.cell(row_sdwan,6,"Bootflash free aligned to current SDD for SDWAN")
		ws_sdwan.cell(row_sdwan,7,"Rommon version")
		ws_sdwan.cell(row_sdwan,8,"Rommon aligned to current SDD for SDWAN")
		ws_sdwan.cell(row_sdwan,9,"Recommended Rommon version for HW for SDWAN 17.2")
		ws_sdwan.cell(row_sdwan,10,"RAM (GB)")
		ws_sdwan.cell(row_sdwan,11,"RAM aligned to current SDD for SDWAN")
		ws_sdwan.cell(row_sdwan,12,"Bootflash total from show version (GB)")
		ws_sdwan.cell(row_sdwan,13,"Bootflash total space ok for SDWAN")
		ws_sdwan.cell(row_sdwan,14,"HSRP interfaces")
		ws_sdwan.cell(row_sdwan,15,"Chassis")
		ws_sdwan.cell(row_sdwan,16,"Chassis support for SDWAN")
		ws_sdwan.cell(row_sdwan,17,"LC(s)")
		ws_sdwan.cell(row_sdwan,18,"LC(s) support for SDWAN")
		ws_sdwan.cell(row_sdwan,19,"Max BW on device(MBPS)")
		ws_sdwan.cell(row_sdwan,20,"Recommended Chassis for SDWAN")	
		ws_sdwan.cell(row_sdwan,21,"Recommended Chassis match existing HW")	
		ws_sdwan.cell(row_sdwan,22,"IGP")
		ws_sdwan.cell(row_sdwan,23,"IGP supported for SDWAN")
		ws_sdwan.cell(row_sdwan,24,"ZBFW zones")
		ws_sdwan.cell(row_sdwan,25,"Is Voice CLI present ?")
		ws_sdwan.cell(row_sdwan,26,"Are Voice cards present ?")
		ws_sdwan.cell(row_sdwan,27,"Are LTE cards present ?")
		ws_sdwan.cell(row_sdwan,28,"Are ATM cards present ?")
		ws_sdwan.cell(row_sdwan,29,"Is device functionaing as voice gateway ?")
	write_cell(ws_sdwan, row_sdwan+1, 1, country)
	write_cell(ws_sdwan, row_sdwan+1, 2, wrk_loc)
	write_cell(ws_sdwan, row_sdwan+1, 3, hn)
	#write_cell(ws_sdwan, row_sdwan+1, 19, str(dev_max_bw))
	write_cell(ws_sdwan, row_sdwan+1, 22, list_to_str(np.unique(np.array(route_protocols))))
	if "RIP" in route_protocols:
		write_cell(ws_sdwan, row_sdwan+1, 23, False)
	else:
		write_cell(ws_sdwan, row_sdwan+1, 23, True)
	write_cell(ws_sdwan, row_sdwan+1, 24, list_to_str(zbfw_list))
	write_cell(ws_sdwan, row_sdwan+1, 25, voice_lines)
	
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 1 , country)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 2 , wrk_loc)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 60, cluster)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 61, wrk_loc_code)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 62, site_name)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 63, region)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 64, dev_ip)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 65, dev_domain)

	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 4, hn)
	if "agi" in hn or "AGI" in hn or "Agi" in hn:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 3,"Agile only router, not a candidate for SDWAN migration")
	#write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 19, str(dev_max_bw))
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 38, list_to_str(np.unique(np.array(route_protocols))))
	if "RIP" in route_protocols or "EIGRP" in route_protocols: #EIGRP now supported in SDWAN, but Unilever wants to standardise OSPF as their core protocol
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 39, False)
	else:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 39, True)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 44, list_to_str(zbfw_list))
	if zbfw_list and (set(zbfw_list).issubset(set(["trusted","internet","guest"]))): 
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 45, True)
	else:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 45, False)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 46, voice_lines)
	
	
	#WAN assessment

	if row_wan == 1:
		ws_wan.cell(row_wan,1,"Country")
		ws_wan.cell(row_wan,2,"Work Location")
		ws_wan.cell(row_wan,3,"Hostname")
		ws_wan.cell(row_wan,4,"Circuit type(s)")
		ws_wan.cell(row_wan,5,"Static route defined in device")
		ws_wan.cell(row_wan,6,"Static route and IP SLA config")
		ws_wan.cell(row_wan,7,"route for default route 0.0.0.0")
		ws_wan.cell(row_wan,8,"show ip route 0.0.0.0")
		ws_wan.cell(row_wan,9,"DEFAULT STATIC 0.0.0.0 rechability")
		ws_wan.cell(row_wan,10,"Route for 145.55.0.0")
		ws_wan.cell(row_wan,11,"show ip route 145.55.0.0 255.255.0.0")
		ws_wan.cell(row_wan,12,"145.55.0.0 routing table entry")
		ws_wan.cell(row_wan,13,"MGMT 145.55.0.0 rechability")
		ws_wan.cell(row_wan,14,"Prod reachability")
		ws_wan.cell(row_wan,15,"Prod alternate reachability")
		ws_wan.cell(row_wan,16,"Neighbor hostnames")
		ws_wan.cell(row_wan,17,"Neighbor ips")
		ws_wan.cell(row_wan,18,"Neighbor reachability")
		ws_wan.cell(row_wan,19,"Route to the neighbors reachable")
		ws_wan.cell(row_wan,23,"ip sla config proper?")
		ws_wan.cell(row_wan,24,"ip sla config for DMVPN/NON-DMVPN?")
		ws_wan.cell(row_wan,25,"default route tracking is present?")
		ws_wan.cell(row_wan,26,"WAN-interface is used in ip sla?")
	
	write_cell(ws_wan, row_wan+1, 1, country)
	write_cell(ws_wan, row_wan+1, 2, wrk_loc)
	write_cell(ws_wan, row_wan+1, 3, hn)
	write_cell(ws_wan, row_wan+1, 4, list_to_str(np.unique(np.array(dev_type))))
	
	#LC support

	if row_lc == 1:
		ws_lc.cell(row_lc,1,"Country")
		ws_lc.cell(row_lc,2,"Work Location")
		ws_lc.cell(row_lc,3,"Hostname")
		ws_lc.cell(row_lc,4,"Module")
		ws_lc.cell(row_lc,5,"Description")
		ws_lc.cell(row_lc,6,"PID")
		ws_lc.cell(row_lc,7,"Serial Number")
		ws_lc.cell(row_lc,8,"Support for SDWAN")
	
	#write_cell(ws_wan, row_wan+1, 6, static_route_default_lines)
	check1_static_route_advertised_from = False

	check2_static_route_backup_path = False
	if not dict_op_hn:
		write_cell(ws_sdwan, row_sdwan+1, 4, "File not found")
		write_cell(ws_wan, row_wan+1, 5, "File not found")
		
	
	chassis = ""
	dev_voice_cards = []
	dev_atm_cards = []
	dev_lte_cards = []
	corp_ips = ["130.24.4.2","130.24.4.4","130.2.2.2"]
	SAP_1 = "130.24.4.2"
	SAP_2 = "130.24.4.4"
	def_route = "0.0.0.0"
	mgmt_route = "145.55.0.0"	
	SAP_1_route = ""
	SAP_2_route = ""
	def_route_p = ""
	mgmt_route_p = ""
	mgmt_route_p1 = ""
	voice_status = ""
	cert = False
	sno = False
	ip_sla_suc=0
	
	history_ipsla_number=re.findall("\d+" , history_ipsla)
	c_type=""
	for i in int_dict.keys():
		if circuit_det[i]=="DMVPN":
			c_type="DMVPN"
	
			

	#ip sla Config Check
	if c_type=="DMVPN":#DMVPN Circuit identifying should be different-based on c_type.
		print("c_type="+ c_type)
		print("in DMVPN IF")
		track_check_DMVPN=re.findall("\d+",track_lines)
		if "source-interface" in icmp_echo_lines:
			Source_int_icmpecho=re.findall("source-interface (.*)",icmp_echo_lines)#Change-Capture othey type of interfces
			Source_int_icmpecho=Source_int_icmpecho[0].replace("\r" , " ").split(" ")
			print(Source_int_icmpecho)
			
			#Source_ip_icmpecho=set(Source_ip_icmpecho)
			if Source_int_icmpecho:
				if lan_wan_dict[Source_int_icmpecho[0]]=="WAN":#Special variable created for identifying LAN/WAN
					print("WAN interface is sourced")
				else:
					print("WAN interface is not sourced")
			#wan_intf_set=set(wan_intf)
			print(wan_intf)
		if "source-ip" in icmp_echo_lines:
			Source_ip_icmpecho=re.findall("source-ip \d+.\d+.\d+.\d+",icmp_echo_lines)
			print(Source_ip_icmpecho)
			if Source_ip_icmpecho:
				for i in int_dict.keys():
					if lan_wan_dict[i]=="WAN" and str(Source_ip_icmpecho[0].split(" ")[1]) in int_dict[i]:
						write_cell(ws_wan, row_wan+1, 26 , True)
						print("WAN interface is sourced")
						break
					
					
					
			
			
				
			
		#Source_ip_icmpecho=re.findall
		print(type(track_check_DMVPN))
		print(track_check_DMVPN)
		track_check_DMVPN=set((track_check_DMVPN))
		print(track_check_DMVPN)
		
		if "track 1 ip sla 1" in track_lines and "track 2 ip sla 2" in track_lines and "track 3 ip sla 3" in track_lines:
			print("tracks are properly configured")
			ip_sla_suc+=1
		else:
			print("tracks are Missing")
		#count=0
		count=ip_sla_lines.count("ip sla schedule")
		if count!=3:
			print("ip sla schedule is not properly configured")
			
		else:
			print("ip sla schedule is properly configured")
			ip_sla_suc+=1
		if icmp_echo_count==3:
			print("ip sla echo is configured properly")
			ip_sla_suc+=1
		
		if frequency_ipsla_count==3:
			print("ip sla frequency is configured properly")
			ip_sla_suc+=1
		
		print(history_ipsla_number)
		if history_ipsla_count==3 and "25" in history_ipsla_number:
			print("ip sla history is configured properly")
			ip_sla_suc+=1
		print(ip_sla_suc)
		if ip_sla_suc==5:
			write_cell(ws_wan, row_wan+1, 23, True)
		else:
			write_cell(ws_wan, row_wan+1, 23, False)
		write_cell(ws_wan, row_wan+1, 24, "DMVPN")
		if def_track_route:
			write_cell(ws_wan, row_wan+1, 25, True)
			
					
	
	#ip sla without DMVPN config check
	else:
		print("c_type="+ c_type)
		print("NOT IN DMVPN IF")
		write_cell(ws_wan, row_wan+1, 24, "NON-DMVPN")
		track_check_withoutDMVPN=re.findall("\d+",track_lines)
		track_check_withoutDMVPN=set((track_check_withoutDMVPN))
		
		if "101" in track_check_withoutDMVPN and "102" in track_check_withoutDMVPN and "103" in track_check_withoutDMVPN and "100" in track_check_withoutDMVPN:
			print("tracks are properly configured")
			ip_sla_suc+=1
		else:
			print("tracks are Missing")
		
		
		if frequency_ipsla_count==3:
			print("ip sla frequency is configured properly")
			ip_sla_suc+=1
		if "ip sla 101" in ip_sla_lines and "ip sla 102" in ip_sla_lines and "ip sla 103" in ip_sla_lines:
			print("ip sla numbers is configured properly")
			ip_sla_suc+=1
		if icmp_echo_count==3:
			print("ip sla echo is configured properly")
			ip_sla_suc+=1
		if ip_sla_suc==4:
			print(ip_sla_suc)
			write_cell(ws_wan, row_wan+1, 23, True)
		else:
			write_cell(ws_wan, row_wan+1, 23, False)
		if def_track_route:
			write_cell(ws_wan, row_wan+1, 25, True)
		if "source-interface" in icmp_echo_lines:
			Source_int_icmpecho=re.findall("source-interface (.*)",icmp_echo_lines)
			Source_int_icmpecho=Source_int_icmpecho[0].replace("\r" , " ").split(" ")
			print(Source_int_icmpecho)
			#Source_ip_icmpecho=set(Source_ip_icmpecho)
			if Source_int_icmpecho:
				if lan_wan_dict[Source_int_icmpecho[0]]=="WAN":
					print("WAN interface is sourced")
				else:
					print("WAN interface is not sourced")
			#wan_intf_set=set(wan_intf)
			print(wan_intf)
		if "source-ip" in icmp_echo_lines:
			Source_ip_icmpecho=re.findall("source-ip \d+.\d+.\d+.\d+",icmp_echo_lines)
			print(Source_ip_icmpecho)
			
			if Source_ip_icmpecho:
				for i in int_dict.keys():
					if lan_wan_dict[i]=="WAN" and str((Source_ip_icmpecho[0]).split(" ")[1]) in int_dict[i]:
						write_cell(ws_wan, row_wan+1, 26 , True)
						print("WAN interface is sourced")
						break
		
				
	for label in sorted(dict_op_hn.keys()): #sorting to ensure show inventory goes before show rom-monitor to capture chassis information
		content = dict_op_hn[label]
		#print ("Type of content ", str(type(content)))
		if "list" in str(type(content)):
			value = content
			#print ("Type of content ", str(type(value)))
		else:
			value = str(content).split("\n")
			#print ("Type of content ", str(type(value)))
		str_value = '\n'.join(value)
		
		if wan_nei_ip and label == "sir_parsed":
			write_cell(ws_wan, row_wan+1, 16, list_to_str(wan_neighbors))
			write_cell(ws_wan, row_wan+1, 17, list_to_str(wan_nei_ip))
			nei_ip_present = []
			route_to_nei_ip_present = []
			for wan_nei in wan_nei_ip:
				if dict_op_hn[label]:
					#print ("Neighbor IP:", wan_nei)
					#print ("dict_op_hn keys", dict_op_hn[label]["routing_table"].keys())
					if "subnet_tree_routing_table" in dict_op_hn[label].keys() and len(str(dict_op_hn[label]["subnet_tree_routing_table"]))>0:
						if wan_nei in dict_op_hn[label]["subnet_tree_routing_table"]:
							nei_ip_present.append(True)
							next_hop, raw_route_string = dict_op_hn[label]["subnet_tree_routing_table"][wan_nei]
							route_to_nei_ip_present.append(raw_route_string + "\tNext Hops --> "+ list_to_str(next_hop) + "\n!!\n")
							#print ("Route to nei: ",route_to_nei_ip_present)
						else:
							nei_ip_present.append(False)
			write_cell(ws_wan, row_wan+1, 18, list_to_str(nei_ip_present))
			write_cell(ws_wan, row_wan+1, 19, list_to_str(route_to_nei_ip_present))
		
		#voice detection
		
		if label == 'show controllers' and str_value : #‘show controllers’ and look for E1 or T1
			if "T1 [\d\/]+ is " in str_value or "E1 [\d\/]+ is " in str_value:  #T1 0/1/0 is up.
				voice_status += "E1 or T1 cards exists\n"
			
		if label == 'show dspfarm' and str_value: #‘show dspfarm’ < If not empty, DSP farm config exist on the routers (DSP is used in E1/T1/FXS/FXO cards)
			if "Invalid input detected" not in str_value and "Incomplete command" not in str_value:
				voice_status += "DSP farm config exists on the routers"	
				if "ADMIN_STATE_DOWN" in str_value:
					voice_status+=" In Admin down state"
			voice_status+="\n"	
				
		if label == 'show voice port summary' and str_value: #‘show voice port summary’ < If output exists, an analog voice port exists such as FXS, FXO card
			if "Invalid input detected" not in str_value and "Incomplete command" not in str_value and ("fxs" in str_value or "fxo" in str_value):
				voice_status += "An analog voice port exists such as FXS, FXO card\n"
			
		if label=='show sip-ua status' and str_value: #‘show sip-ua status’ < shows if we’re being a SIP user agent. Grep for ‘ENABLED’. ‘sip-ua’ will also be present in config.
			if "Invalid input detected" not in str_value and ('ENABLED' in str_value and 'SIP-UA' in str_value):
				voice_status += "Acting as SIP user agent\n"	
			
		if label == 'show sip register status' and str_value: #‘show sip register status’ < Shows if we’re registered to a SIP UA/SIP server. Will have line number and registered time.
			if "Invalid input detected" not in str_value and "Registrar is not configured" not in str_value:
				if "Tacacs session has expired.Please re-login to continue" in str_value:
					voice_status += "SIP Registrar TACACS session expired, login to continue\n"	
				else:
					voice_status += "Configured as SIP Registrar\n"	
		#end of voice		
		
		
		#WAN failover
		if label == 'show run | in ip route': #Outliers  - Default route addition of IP SLA
			if "ip route 0.0.0.0 0.0.0.0" in str_value:
				write_cell(ws_wan, row_wan+1, 5, True)
			else:
				write_cell(ws_wan, row_wan+1, 5, False)
			#write_cell(ws_wan, row_wan+1, 6, str_value)
		if def_route in label:
			if label == "show ip route "+def_route:
				known_via = re.search("Known via \"(.*)\"",str_value)
				def_route_p+=sh_ip_route(str_value,def_route)
				if known_via:
					write_cell(ws_wan, row_wan+1, 7, True)
				else:
					write_cell(ws_wan, row_wan+1, 7, False)
				write_cell(ws_wan, row_wan+1, 8, str_value)
				if static_route_default_lines.count("ip route 0.0.0.0 0.0.0.0")>1:
					def_route_p += "More than one default static route on device"
				if static_route_default_lines.count("ip route vrf DMVPN-TRANSPORT-INTERNET 0.0.0.0 0.0.0.0")>1:
					def_route_p += "More than one VRF static route on device"				
			if label == "show ip bgp "+def_route:
				if "entry for 0.0.0.0/0" in str_value:
					def_route_p += "Via default route in BGP\n"
			if label == "show ip eigrp topology "+def_route:
				def_route_p += eigrp_topology_route(str_value,def_route)
			if label == "show ip ospf database external "+def_route:
				def_route_p += ospf_database_route(str_value,def_route)

		if mgmt_route in label:
			if label == 'show ip route 145.55.0.0 255.255.0.0':
				known_via = re.search("nown via \"(.*)\"", str_value)
				mgmt_route_p1 += sh_ip_route(str_value,mgmt_route)
				if known_via:
					write_cell(ws_wan, row_wan+1, 10, True)
				else:
					write_cell(ws_wan, row_wan+1, 10, False)
				write_cell(ws_wan, row_wan+1, 11, str_value)
			if label == "show ip bgp "+mgmt_route:
				mgmt_route_p += ip_bgp_route(str_value,mgmt_route)
			if label == "show ip eigrp topology "+mgmt_route:
				mgmt_route_p += eigrp_topology_route(str_value,mgmt_route)
			if label == "show ip ospf database external "+mgmt_route:
				mgmt_route_p += ospf_database_route(str_value,mgmt_route)

		if SAP_1 in label: #Outliers  - Change to include all IPs for corporate as a list function
			if label == "show ip route "+SAP_1:
				SAP_1_route += sh_ip_route(str_value,SAP_1)
			if label == "show ip bgp "+SAP_1:
				SAP_1_route += ip_bgp_route(str_value,SAP_1)
			if label == "show ip eigrp topology "+SAP_1:
				SAP_1_route += eigrp_topology_route(str_value,SAP_1)
			if label == "show ip ospf database external "+SAP_1:
				SAP_1_route += ospf_database_route(str_value,SAP_1)
		if SAP_2 in label:
			if label == "show ip route "+SAP_2:
				SAP_2_route += sh_ip_route(str_value,SAP_2)
			if label == "show ip bgp "+SAP_2:
				SAP_2_route += ip_bgp_route(str_value,SAP_2)
			if label == "show ip eigrp topology "+SAP_2:
				SAP_2_route += eigrp_topology_route(str_value,SAP_2)
			if label == "show ip ospf database external "+SAP_2:
				SAP_2_route += ospf_database_route(str_value,SAP_2)
				
		#WAN failover section ends
		
		if label == 'show inventory':
			#print ("show inventory\n")
			line_number = 1
			item_dict={}
			lc = []
			lc_support = []
			lc_count = 0
			for line_item in value:
				#print (" Line : + "+line_item)
				if line_number == 1:
				    results = re.match("NAME: \"(.*)\", DESCR: \"(.*)\"", line_item)
				    serial = pid = vid = descr = slot = ""
				    if results:
				    	#print " Results:"
				    	name = results.group(1)
				    	descr = results.group(2)
				    	line_number = 2 
				if line_number == 2:                      
				    pid_details = re.match("PID: ([^\s]*)\s*,\s*VID: ([^\s]*)\s*, SN: (.*)",line_item)
				    if pid_details:
				    	#print ("PID", pid_details.group(1)) 
				    	pid = pid_details.group(1)
				    	vid = pid_details.group(2)
				    	serial = pid_details.group(3)
				    	#print pid, vid, serial
				    	line_number = 1
				    	if "Chassis" in name or "chassis" in descr or "CHASSIS" in name or "CHASSIS" in descr:
				    		#print ("Chassis in name")
				    		write_cell(ws_sdwan, row_sdwan+1, 15, pid)
				    		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 16, pid)
				    		if not device_type:
				    			write_cell(ws,row+1,22,pid)
				    		chassis = pid
				    		if chassis in chassis_bw_map.keys():
				    			#print (dev_max_bw, "MB,", chassis_bw_map[chassis])
				    			lc_count +=1
				    			write_cell(ws_lc, row_lc+lc_count, 1, country)
				    			write_cell(ws_lc, row_lc+lc_count, 2, wrk_loc)
				    			write_cell(ws_lc, row_lc+lc_count, 3, hn) 
				    			write_cell(ws_lc, row_lc+lc_count, 4, name)
				    			write_cell(ws_lc, row_lc+lc_count, 5, descr) 
				    			write_cell(ws_lc, row_lc+lc_count, 6, pid)
				    			write_cell(ws_lc, row_lc+lc_count, 7, serial)
				    			if int(dev_max_bw) <= int(chassis_bw_map[chassis]["to"]):
				    				#print ("matched dev_max_bw")
				    				write_cell(ws_sdwan, row_sdwan+1, 21, True) 
				    				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 25, True) 				
				    			else:
				    				#print ("didnt match dev_max_bw")
				    				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 25, False)
				    		else:
				    			write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 25, False)
				    		for c_key in chassis_bw_map.keys():
				    			#print (dev_max_bw,"MB, ",chassis_bw_map[c_key])
				    			if (int(dev_max_bw) >= int(chassis_bw_map[c_key]["from"])) and (int(dev_max_bw) < int(chassis_bw_map[c_key]["to"]) ):
				    				write_cell(ws_sdwan, row_sdwan+1, 20, str(c_key))
				    				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 24, str(c_key))
				    				#print ("Suggested chassis by BW:",c_key)
				    			
				    		if pid in sdwan_supported_hw:
				    			write_cell(ws_sdwan, row_sdwan+1, 16, True)
				    			write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 17, True)
				    			write_cell(ws_lc, row_lc+lc_count, 8, True) #for chassis support for SDWAN
				    		else:
				    			write_cell(ws_sdwan, row_sdwan+1, 16, False)
				    			write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 17, False)
				    			write_cell(ws_lc, row_lc+lc_count, 8, False)
				    	if "subslot" in name and "transceiver" not in name and "Front Panel" not in descr and "BUILT-IN" not in pid :
				    		if not re.search("(\w+)",serial):
				    			continue
				    		lc_count +=1
				    		write_cell(ws_lc, row_lc+lc_count, 1, country)
				    		write_cell(ws_lc, row_lc+lc_count, 2, wrk_loc)
				    		write_cell(ws_lc, row_lc+lc_count, 3, hn) 
				    		write_cell(ws_lc, row_lc+lc_count, 4, name)
				    		write_cell(ws_lc, row_lc+lc_count, 5, descr) 
				    		write_cell(ws_lc, row_lc+lc_count, 6, pid)
				    		write_cell(ws_lc, row_lc+lc_count, 7, serial)
				    		lc.append(pid)
				    		#print ("inside subslot", name,descr,pid,"Serial:",serial)
				    		if pid in sdwan_isr4k_lc_support and "ISR4" in chassis:
				    			lc_support.append(True)
				    			write_cell(ws_lc, row_lc+lc_count, 8, True)
				    		elif "ISR4" in chassis:
				    			lc_support.append(False)
				    			write_cell(ws_lc, row_lc+lc_count, 8, False)
				    		else:
				    			lc_support.append("Not recognised")
				    			write_cell(ws_lc, row_lc+1, 8, "Not recognised")
				    	
				    	if pid in voice_cards:
				    		dev_voice_cards.append(pid)
				    	if pid in lte_cards:
				    		dev_lte_cards.append(pid)
				    	if pid in atm_cards:
				    		dev_atm_cards.append(pid)
				    	serial = pid = vid = descr = name = ""
			write_cell(ws_sdwan, row_sdwan+1, 17, list_to_str(lc))
			if lc:
				write_cell(ws_sdwan, row_sdwan+1, 18, list_to_str(lc_support))
			else:
				write_cell(ws_sdwan, row_sdwan+1, 18, True)
			write_cell(ws_sdwan, row_sdwan+1, 26, list_to_str(dev_voice_cards))
			write_cell(ws_sdwan, row_sdwan+1, 28, list_to_str(dev_lte_cards))
			write_cell(ws_sdwan, row_sdwan+1, 29, list_to_str(dev_atm_cards))
			
			write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 18,  list_to_str(lc))
			if lc:
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 20, list_to_str(lc_support))
				if "False" in lc_support:
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 19, False)
				else:
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 19, True)
			else:
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 20, True)
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 19, True)
			write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 47, list_to_str(dev_voice_cards))
			write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 50, list_to_str(dev_lte_cards))
			write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 51, list_to_str(dev_atm_cards))
		if label == 'show standby brief':
			if value:
				#print ("standby: ", value, "\n\n", str_value)
				interfaces = re.findall("\n\s*([\w\.\/\d]+)",str_value)
				if interfaces and 'list' in str(type(interfaces)):
					#print (interfaces)
					if len(interfaces)>0:
						if "Interface" in interfaces:
							interfaces.remove("Interface")
						write_cell(ws_sdwan, row_sdwan+1, 14, list_to_str(interfaces))
						if len(interfaces)>0:
							write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 42, False)
							write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 43, list_to_str(interfaces))
						else:
							write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 42, False)
							write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 43, "Configured, but no interfaces active")	
					else:
						write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 42, True)	
				else:
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 42, True)
			else:
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 42, True)
		if label == 'show version':
			RAM = re.search("(\d+)K bytes of physical memory",str(value))
			boot = re.search("(\d+)K bytes of flash memory at bootflash",str(value))
			version = re.search("Version ([\d\.\S]+)\s*\n",str(value))
			if version: 
				write_cell(ws,row+1,23,version.group(1))
			if RAM:
				#print ("Found ram in version")
				ram_space = int(RAM.group(1))/1000000
				write_cell(ws_sdwan, row_sdwan+1, 10, str(round(ram_space,2)))
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 30, str(round(ram_space,2)))
				if ram_space > 13:
					write_cell(ws_sdwan, row_sdwan+1, 11, True)
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 31, True)
				else:
					write_cell(ws_sdwan, row_sdwan+1, 11, False)
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 31, False)
			else:
				write_cell(ws_sdwan, row_sdwan+1, 11, False)
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 31, False)
			if boot:
				#print ("found boot in version")
				boot_space = int(boot.group(1))/1000000
				write_cell(ws_sdwan, row_sdwan+1, 12, str(round(boot_space,2)))
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 26, str(round(boot_space,2)))
				if boot_space > 13:
					write_cell(ws_sdwan, row_sdwan+1, 13, True)
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 27, True)
				else:
					write_cell(ws_sdwan, row_sdwan+1, 13, False)
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 27, False)
			else:
				write_cell(ws_sdwan, row_sdwan+1, 13, False)
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 27, False)
		if label == 'dir bootflash:':
			space_avl = re.search("(\d+) bytes total \((\d+) bytes free\)",str(value))
			if space_avl:
				#print ("Found spaces ",space_avl.groups())
				bootflash_free_space = int(space_avl.group(2))/1000000000
				write_cell(ws_sdwan, row_sdwan+1, 4, str(round((int(space_avl.group(1))/1000000000),2)))
				write_cell(ws_sdwan, row_sdwan+1, 5, str(round(bootflash_free_space,2)))
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 28, str(round(bootflash_free_space,2)))
				if bootflash_free_space >=1.5:
					write_cell(ws_sdwan, row_sdwan+1, 6, True)
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 29, True)
				else:
					write_cell(ws_sdwan, row_sdwan+1, 6, False)
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 29, False)
			else:
				write_cell(ws_sdwan, row_sdwan+1, 6, False)
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 29, False)
		if label == 'show rom-monitor 0':
			rom_ver = re.search("Version ((\d+)\.(\d+)(\(\dr\)))",str(value))
			dev_rom_val = ""
			if rom_ver:
				dev_rom_val = str(rom_ver.group(1))
				write_cell(ws_sdwan, row_sdwan+1, 7, dev_rom_val)
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 35, dev_rom_val)
			if chassis in sdwan_ver_17_2.keys() and dev_rom_val:
				rom_val = sdwan_ver_17_2[chassis]
				rv1 = re.search("(\d+)\.(\d+)\((\d+)(r?)\)",rom_val)
				rv2 = re.search("(\d+)\.(\d+)\((\d+)(r?)\)",dev_rom_val)
				if rom_val == dev_rom_val:
					write_cell(ws_sdwan, row_sdwan+1, 8, True)
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 36, True)
				elif rv1 and rv2:
					#print (rv1.groups())
					#print (rv2.groups())
					if int(rv2.group(1))>=int(rv1.group(1)) and int(rv2.group(2))>=int(rv1.group(2)) and int(rv2.group(3))>=int(rv1.group(3)) and rv2.group(4)==rv1.group(4):
						#print ("rommon fine")
						write_cell(ws_sdwan, row_sdwan+1, 8, True)
						write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 36, True)
					else:
						#print ("Need rommom upgrade")
						write_cell(ws_sdwan, row_sdwan+1, 8, False)
						write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 36, False)
				else:
					#print ("Need rommom upgrade")
					write_cell(ws_sdwan, row_sdwan+1, 8, False)
					write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 36, False)
				write_cell(ws_sdwan, row_sdwan+1, 9, rom_val)
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 37, rom_val)	
			else:
				write_cell(ws_sdwan, row_sdwan+1, 8, False)
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 36, False)
				write_cell(ws_sdwan, row_sdwan+1, 9, "Device does not support SDWAN, needs upgrade")
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 37, "Device does not support SDWAN, needs upgrade")	
		if label == "show license udi":
			for line_lic in value:
				#print ("License line ", line_lic)
				if "SlotID" not in line_lic and "------" not in line_lic:
					sno_udi = re.search("\*\s+\S+\s+(\S+)\s+([\S\:]+)",line_lic)
					if sno_udi:
						#print (sno_udi.groups())
						write_cell(ws, row+1, 31, sno_udi.group(1))
						write_cell(ws, row+1, 32, sno_udi.group(2))
						if not sno:
							write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 53, sno_udi.group(2).split(":")[0])
							write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 54, sno_udi.group(1))
							write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 55, False)	
						
						
						
		if label == "show crypto pki certificates CISCO_IDEVID_SUDI":
			certno = re.findall(" Certificate Serial Number \(hex\): (\S+)",str_value)
			sno_pid = re.search("Serial Number: PID:([\w\/]+) SN:(\w+)",str_value)
			if certno and 'list' in str(type(certno)):
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 52, certno[0])
				cert = True
			if sno_pid:
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 53, sno_pid.group(1))
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 54, sno_pid.group(2))	
				sno = True
			if cert and sno:
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 55, True)
			else:
				write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 55, False)

	# outside for loop for dict_op_hn
	if chassis in sdwan_ver_17_2.keys():
		rom_val = sdwan_ver_17_2[chassis]
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 37, rom_val) #24-sep - Check this 
	write_cell(ws_wan, row_wan+1, 14, SAP_1_route)
	write_cell(ws_wan, row_wan+1, 15, SAP_2_route)
	write_cell(ws_wan, row_wan+1, 9, def_route_p)
	write_cell(ws_wan, row_wan+1, 12, mgmt_route_p1)
	write_cell(ws_wan, row_wan+1, 13, mgmt_route_p)
	write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 48, voice_status)
	if voice_status or dev_voice_cards or voice_lines:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 49, False)
	else:
		write_cell(ws_sdwan_rtr, row_sdwan_rtr+1, 49, True)
	write_cell(ws_sdwan, row_sdwan+1, 29, voice_status)
	return dict_op_hn

#Read Assessment Sheet
def readAssessmentFile(filename):
	rawdata = pd.read_excel(filename, header=0)
	reqdData = rawdata[['Master_RFP Reference', 'RFP_REF', 'Mgmt_IP', 'Site Name', 'Circuit Identifier', 'Router Hostname','Old CKT', 'Supplier']]
	##print(reqdData.head())
	return reqdData
	
def readMasterTracker(filename):
	rawdata = pd.read_excel(filename, header=0)
	reqdData = rawdata[["New Circuit Reference","Router Hostname","New Router Hostname","Ordered Bandwidth (Mbps)","IP Address Ordered"]]
	##print(reqdData.head())
	return reqdData 
	
def readCktSheet(filename):
	reqdData = pd.read_excel(filename, sheet_name='Circuit Info List', usecols=['Site Name','Circuit Identifier','Circuit Type','Circuit BW','Router (CISCO) Host Name','Router (CISCO) IP Address'])
	##print(reqdData.head())
	return reqdData

def readSplunkSheet(filename):
	rawData = pd.read_excel(filename, header =0, sheet_name='Sheet2' ) #sheet2 for all devices, sheet6 for 180 priority sites
	#reqdData = rawData[['Region','Site','Device IP','Device Name and Domain','Device Name']]
	reqdData = rawData[['Cluster','Country','Work Location','Work Location Code','CMSP - Customer Name','CMSP - Region','CMSP - Site','Device Onboarding Date','Device ID','Asset ID','Device IP','Device Name & Domain','Device Category','Model','Serial','IOS','Device State','Impact','Device Name','Region:Site']]
	#print(reqdData.head())
	return reqdData

def readCoralTaskOp(filename):
	rawData = pd.read_excel(filename, header=0 )
	reqdData = rawData[['IP','Host Name','Status','Error','show run | in ip route','show ip route 0.0.0.0','show ip bgp 0.0.0.0','show ip eigrp topology 0.0.0.0','sh ip ospf database external 0.0.0.0','show cdp nei','show ip ospf nei','show ip eigrp nei','show ip bgp nei','show run | s router ospf','show run | s router rip','show run | s router eigrp','show run | s router bgp','show ip route 145.55.0.0 255.255.0.0','show ip eigrp topology 145.55.0.0','sh ip ospf database external 145.55.0.0','show configuration | in hostname','show ip int br | in Loopback','show ip int br | in Tu','show process cpu | in CPU','show process memory | in Used','show inventory | in NAME','show version | in IOS','show ip bgp summary | in local AS','show cdp neighbor detail | in Device','dir nvram:','dir bootflash:','show rom-monitor 0','show ip route','show ip route vrf DMVPN-TRANSPORT-INTERNET']]
	i = 0
	for value in reqdData['Host Name']:
		reqdData['Host Name'][i] = str(reqdData['Host Name'][i]).split(".")[0]
		i = i+ 1
	#print (reqdData['Host Name'])
		
	return reqdData

#Freeze panes at B2 for each sheet in wb

def freeze_panes(wb,freeze_cell):
	for sheet in wb.worksheets:
		c = sheet[freeze_cell]
		sheet.freeze_panes = c
		for cell in sheet[1]:
			cell.style = highlight
			cell.alignment =  Alignment(wrap_text=True,vertical='top') 
		#if "SDWAN" in sheet.name:
		#	for cell in sheet[2]:
		#		cell.style = highlight
		#		cell.alignment =  Alignment(wrap_text=True,vertical='top') 
		for col in sheet.columns:
			col_name = re.findall('\w\d', str(col[0]))
			col_name = col_name[0]
			col_name = re.findall('\w', str(col_name))[0]
			sheet.column_dimensions[col_name].width = 20
		sheet.row_dimensions[1].height = 40
		

def run_nw_chk(ws_chk,ckt_sheet,Read_Excel):

	row_chk = ws_chk.max_row
	if row_chk == 1:
		ws_chk.cell(row_chk,1,"Site Name")
		ws_chk.cell(row_chk,2,"Circuit ID given")
		ws_chk.cell(row_chk,3,"Circuit Type given")
		ws_chk.cell(row_chk,4,"Circuit BW given")
		ws_chk.cell(row_chk,5,"Router hostname given")
		ws_chk.cell(row_chk,6,"Router IP given")
		ws_chk.cell(row_chk,7,"Circuit ID seen")
		ws_chk.cell(row_chk,8,"Circuit Type seen")
		ws_chk.cell(row_chk,9,"Circuit BW seen")
		ws_chk.cell(row_chk,10,"Router hostname seen")
		ws_chk.cell(row_chk,11,"Router IP seen")
		ws_chk.cell(row_chk,12,"ZBFW seen")
		ws_chk.cell(row_chk,13,"NAT pool seen")
		ws_chk.cell(row_chk,14,"Network mask on WAN edge interface")
		ws_chk.cell(row_chk,15,"IP SLA config seen")
		ws_chk.cell(row_chk,16,"Routing protocols seen")
		ws_chk.cell(row_chk,17,"Image")
		ws_chk.cell(row_chk,18,"WAN interface")
		ws_chk.cell(row_chk,19,"Device Type")
		ws_chk.cell(row_chk,20,"Region")
		ws_chk.cell(row_chk,21,"Site Name")
		
	
	for i in range(len(ckt_sheet)):
		row_chk = ws_chk.max_row
		ws_chk.cell(row_chk+1,1,ckt_sheet['Site Name'][i])
		ws_chk.cell(row_chk+1,2,ckt_sheet['Circuit Identifier'][i])
		ws_chk.cell(row_chk+1,3,ckt_sheet['Circuit Type'][i])
		ws_chk.cell(row_chk+1,4,ckt_sheet['Circuit BW'][i])
		ws_chk.cell(row_chk+1,5,ckt_sheet['Router (CISCO) Host Name'][i])
		ws_chk.cell(row_chk+1,6,ckt_sheet['Router (CISCO) IP Address'][i])
		try:
			hostname = ckt_sheet['Router (CISCO) Host Name'][i]
			REF = Read_Excel.loc[Read_Excel['Hostname'] == hostname]
			ws_chk.cell(row_chk+1,7,list_to_str(REF['Circuit ID seen'].dropna()))
			ws_chk.cell(row_chk+1,8,list_to_str(REF['Device Type'].dropna()))
			ws_chk.cell(row_chk+1,9,list_to_str(REF['Bandwidth of physcial interfaces'].dropna()))
			ws_chk.cell(row_chk+1,10,hostname)
			ws_chk.cell(row_chk+1,11,list_to_str(REF['IP'].dropna()))
			ws_chk.cell(row_chk+1,12,list_to_str(REF['ZBFW'].dropna()))
			ws_chk.cell(row_chk+1,13,list_to_str(REF['NAT Mask in CIDR'].dropna()))
			ws_chk.cell(row_chk+1,14,list_to_str(REF['Existing net mask on WAN interfaces'].dropna()))
			ws_chk.cell(row_chk+1,15,list_to_str(REF['IP SLA'].dropna()))
			ws_chk.cell(row_chk+1,16,list_to_str(REF['Routing in use'].dropna()))
			ws_chk.cell(row_chk+1,17,list_to_str(REF['Image'].dropna()))
			ws_chk.cell(row_chk+1,18,list_to_str(REF['WAN interface'].dropna()))
			ws_chk.cell(row_chk+1,19,list_to_str(REF['Device Type'].dropna()))
			ws_chk.cell(row_chk+1,20,list_to_str(REF['Region'].dropna()))
			ws_chk.cell(row_chk+1,21,list_to_str(REF['Site Name'].dropna()))
			print ("Matching " + str(hostname) +"\n")
		except:
			print ("Hostname not found in uploaded files :", str(hostname))
			
	#Loop for checking for extra hostnames
	for i in range(len(Read_Excel)):
		row_chk = ws_chk.max_row
		if 	(Read_Excel['Hostname'][i] not in ckt_sheet['Router (CISCO) Host Name'].unique()):
			hostname = Read_Excel['Hostname'][i]
			REF = Read_Excel.loc[Read_Excel['Hostname'] == hostname]
			ws_chk.cell(row_chk+1,7,list_to_str(REF['Circuit ID seen'].dropna()))
			ws_chk.cell(row_chk+1,8,list_to_str(REF['Device Type'].dropna()))
			ws_chk.cell(row_chk+1,9,list_to_str(REF['Bandwidth of physcial interfaces'].dropna()))
			ws_chk.cell(row_chk+1,10,hostname)
			ws_chk.cell(row_chk+1,11,list_to_str(REF['IP'].dropna()))
			ws_chk.cell(row_chk+1,12,list_to_str(REF['ZBFW'].dropna()))
			ws_chk.cell(row_chk+1,13,list_to_str(REF['NAT Mask in CIDR'].dropna()))
			ws_chk.cell(row_chk+1,14,list_to_str(REF['Existing net mask on WAN interfaces'].dropna()))
			ws_chk.cell(row_chk+1,15,list_to_str(REF['IP SLA'].dropna()))
			ws_chk.cell(row_chk+1,16,list_to_str(REF['Routing in use'].dropna()))
			ws_chk.cell(row_chk+1,17,list_to_str(REF['Image'].dropna()))
			ws_chk.cell(row_chk+1,18,list_to_str(REF['WAN interface'].dropna()))
			ws_chk.cell(row_chk+1,19,list_to_str(REF['Device Type'].dropna()))
			ws_chk.cell(row_chk+1,20,list_to_str(REF['Region'].dropna()))
			ws_chk.cell(row_chk+1,21,list_to_str(REF['Site Name'].dropna()))
			print ("Unmatched " + str(hostname) +"\n")

def readFileToDict(filename):
	print ("Creating dictionary of manually collected show commands")
	file_contents = open(filename,"r")
	hostname = ""
	dict_op = {}
	show_command = ""
	
	for line in file_contents:
		tl0 = re.search("^(.*)\#(term.* len.* 0)",line)
		sh_cmd = re.search("^(.*)\#(sh.*)",line)
		dir_cmd = re.search("^(.*)\#(dir.*)",line)
		if tl0:
			#if hostname != "" and hostname!= tl0.group(1):
				#print ("End of hostname ", hostname)	
				
			hostname = tl0.group(1)
			dict_op[hostname] = {}
			show_command = ""
			#print ("Start of hostname ", hostname)
			
		if sh_cmd:
			show_command = sh_cmd.group(2)
			if hostname == sh_cmd.group(1):
				#print ("show command ",show_command," found")
				dict_op[hostname][show_command] = line
		if dir_cmd:
			show_command = dir_cmd.group(2)
			if hostname == dir_cmd.group(1):
				dict_op[hostname][show_command] = line
		if show_command:
			dict_op[hostname][show_command] += line
		
	#print  (dict_op)
	return dict_op
	
def readCoralFileToDict(filename):
	print ("Creating dictionary of file from CMS Coral collected show commands")
	file_contents = open(filename,"r")
	op_folder = Path(filename).parent
	hostname = ""
	dict_op = {}
	show_command = ""
	host = False
	
	for line in file_contents:
		ip = re.search("\[([\d\.]+)\]",line)
		sh_cmd = re.search("^\((show .*)\)", line)
		dir_cmd = re.search("^\((dir .*)\)", line)
		if ip:
			ip_n = ip.group(1)
		if sh_cmd and "hostname" in line:
			host = True
			show_command = ""
			continue
		if host:
			if hostname and ip_n:
				json_object = json.dumps(dict_op[hostname])
				with open(os.path.join(op_folder,hostname+"_"+ip_n+".json"), "w") as outfile: 
					outfile.write(json_object) 	
				print ("JSON file created ", hostname)
				hostname = ""
			hn = re.search("hostname (.*)", line)
			if hn:
				hostname = hn.group(1)
				host = False
				dict_op[hostname] = {}
				show_command = ""
			#print ("Start of hostname ", hostname)
			
		if sh_cmd:
			show_command = sh_cmd.group(1)
			dict_op[hostname][show_command] = ""
		if dir_cmd:
			show_command = dir_cmd.group(1)
			dict_op[hostname][show_command] = ""
		if show_command:
			dict_op[hostname][show_command] += line
		
	#print  (dict_op)
	return dict_op

def populate_json_dict(json_folder):
	print ("Creating dictionary of CORAL task from JSON files")
	dict_op = {}
	hostname = ""
	show_command = ""
	unknown_hostname_count = 0	
	for f in os.listdir(json_folder):
		if ".json" in f:
			#print ("Reading JSON folder data for ", f)
			with open(os.path.join(json_folder,f)) as json_file:
				#print (f)
				json_data = json.load(json_file)
			if json_data != {}:
				if "show configuration | in hostname" in json_data.keys():
					host = re.search("hostname (\S+)",str(json_data["show configuration | in hostname"]))
					if host:
						hn = host.group(1)
				else:
					host = re.search("^\.(\d+\.\d+\.\d+\.\d+)",f)
					unknown_hostname_count += 1
					if host:
						hn = host.group(1)
					else:
						hn = "unknown"+str(unknown_hostname_count)
				if hn:
					dict_op[hn] = json_data
			
			
			
			#file_contents = open(os.path.join(json_folder,f),"r")
			#for line in file_contents:
			#	tl0 = re.search("^(.*)\#(term.* len.* 0)",line)
			#	sh_cmd = re.search("^(.*)\#(sh.*)",line)
			#	dir_cmd = re.search("^(.*)\#(dir.*)",line)
		#
			#	if tl0:	
			#		hostname = tl0.group(1)
			#		dict_op[hostname] = {}
			#		show_command = ""
			#		#print ("Start of hostname ", hostname)
			#		
			#	if sh_cmd:
			#		show_command = sh_cmd.group(2)
			#		if hostname == sh_cmd.group(1):
			#			#print ("show command ",show_command," found")
			#			dict_op[hostname][show_command] = line
			#	if dir_cmd:
			#		show_command = dir_cmd.group(2)
			#		if hostname == dir_cmd.group(1):
			#			dict_op[hostname][show_command] = line
			#	if show_command:
			#		dict_op[hostname][show_command] += line

	return dict_op

def np_to_dict(np_folder):
	print ("Creating dictionary of NP download")
	metadata_folder = os.path.join(np_folder,"MetaData")
	cli_folder = os.path.join(np_folder, "CLI")
	base_folder = os.path.join(np_folder,"Config")
	np_dict = {}
	for f in os.listdir(base_folder):
		if ".zip" or "DS_Store" in f:
			continue
		np_dict[f] = {}
		file_path=open(os.path.join(base_folder,f,f+"-running.txt"),"r")
		np_dict[f]["show run"] = file_path.readlines()
		file_path.close()
		metadata_file = open(os.path.join(metadata_folder,f+".xml"),"r")
		np_dict[f]["metadata"] = metadata_file.readlines()
		metadata_file.close()
		for cli in os.listdir(os.path.join(cli_folder,f)):
			cli_name = re.search("\-(show.*)\.txt",cli)
			if cli_name:
				cli_file = open(os.path.join(cli_folder,f,cli),"r")
				np_dict[f][re.sub("\_"," ",cli_name.group(1))] = cli_file.readlines()
				cli_file.close()
		
	return np_dict
	

#MAIN PROGRAM

#base_folder = "/Users/ashrithaganapathy/Work/PROJECTS/Unilever/Information/BT CONFIGURATIONS FILES!!!/"

#base_folder = "/Users/ashrithaganapathy/Work/PROJECTS/Unilever/Information/Show commands collected/UL Week 1/"


start_time = time.time()
#assessment_sheet = "./Assessment_sheetv0.2-config.xlsx"
#site_dev_ckt_sheet = "./CISCO-ULDevicesandCircuits27.9.2019.xlsx"
#master_tracker = "./Unilever WAN Migration MASTER Tracker.xlsx"
#coral_task_op_excel = "/Users/ashrithaganapathy/Work/PROJECTS/Unilever/CMS_Coral_report/17Jul2020/c5c4325d-a59c-41ee-86b7-c63a8671ae3b.xlsx"

policy_file="/Users/praette/Desktop/environment/UL_NW_Assessment/PM_INTERNET_OUT.txt"
splunk_sheet="/Users/praette/Desktop/environment/UL_NW_Assessment/Script_input/Splunk_OSS_dashboard/DeviceInventorySplunk05102020.xlsx"
data_file_from_nw="/Users/praette/Desktop/environment/UL_NW_Assessment/Script_input/UL-test-wan_priority-6jul-1.txt"
coral_task_json="/Users/praette/Desktop/environment/UL_NW_Assessment/Script_input/CMS_Coral_report/3Sep2020/result_folder/"
np_folder = "/Users/praette/Desktop/environment/UL_NW_Assessment/Script_input/NP_download/RawInventory_Export_2020-Sep-10_08-05-05/"
coralfile = "/Users/praette/Desktop/environment/UL_NW_Assessment/Script_input/CMS_Coral_report/5oct20202-2/16c801c6-204c-4fc5-9111-9f6cbf8a48d0.txt"
np_dict = np_to_dict(np_folder)


#smartsheet master tracker download
access_token = 'ghippwg5xupf7mjpgquc6c9acg'
sheet_id = 988980360570756
ss = smartsheet.Smartsheet(access_token)
ss.errors_as_exceptions(True)
#commenting out the smartsheet download due to internal server error from API. Will uncomment for next run #16-sep-2020
#download_MT = ss.Sheets.get_sheet_as_excel(sheet_id,os.getcwd(),"./Unilever WAN Migration MASTER Tracker.xlsx")
master_tracker = "./Unilever WAN Migration MASTER Tracker.xlsx"

with open(policy_file, 'r') as myfile:
  policy_data = myfile.read()
if not os.path.exists(np_folder):
	print (" NP inventory output folder does not exist")
	
if not os.path.exists(coral_task_json):
	print ("CORAL task output folder does not exist")
	
if not os.path.exists(splunk_sheet):
	print (" Splunk sheet does not exist")

if not os.path.exists(data_file_from_nw):
	print (" Network download file does not exist")
	
old_path=os.getcwd()

#reqdData = readAssessmentFile(assessment_sheet)
masterTracker = readMasterTracker(master_tracker)
#ckt_sheet = readCktSheet(site_dev_ckt_sheet)
splunk_data = readSplunkSheet(splunk_sheet)
#dict_op = readFileToDict(data_file_from_nw)
#coral_op = readCoralTaskOp(coral_task_op_excel)
json_dict = populate_json_dict(coral_task_json)
#print (json_dict)
dict_op = readCoralFileToDict(coralfile)

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=15)
bd = Side(style='thick', color="111111")
highlight.border = Border(bottom=bd) #left=bd, top=bd, right=bd, 
yellowFill = PatternFill(start_color='FFEE08',   # - yellow
				   end_color='FFEE08',
				   fill_type='solid')
redFill = PatternFill(start_color='FFFF0000',   # - red
				   end_color='FFFF0000',
				   fill_type='solid')
blueFill = PatternFill(start_color='000000FF',   # - blue
				   end_color='000000FF',
				   fill_type='solid')
output_path = "Script_output/UL_Read"
op_dir = os.path.join(old_path,output_path)
if not os.path.exists(op_dir):
	os.makedirs(op_dir)

store_dict={}
wb = Workbook()

ws_sdwan_rtr = wb.create_sheet ("SDWAN Router Readiness summary")
ws_summ = wb.create_sheet("WAN Assessment summary")
ws = wb.create_sheet("RTR summary")
ws_int = wb.create_sheet("Interface summary")
ws_rtng = wb.create_sheet("Routing summary")
ws_plcy = wb.create_sheet("Policy summary")
ws_cdp = wb.create_sheet("P2P Neighbor")
ws_wan = wb.create_sheet ("WAN Failover Assessment")
ws_routes = wb.create_sheet ("Per Site Route Comparison")
ws_dhcp = wb.create_sheet ("DHCP summary")
ws_sdwan = wb.create_sheet ("SDWAN Readiness summary")
ws_sdwan_site = wb.create_sheet ("SDWAN Device Readiness score")
ws_lc = wb.create_sheet ("Module support for SDWAN")

	

wb1=Workbook()
ws_chk = wb1.create_sheet("Compare NW to CKT data given")
wb.remove(wb['Sheet'])
wb1.remove(wb1['Sheet'])

wb2=Workbook()

wb.add_named_style(highlight)
wb2.add_named_style(highlight)
hostcount = 0

dict_all_hn = {}
sir_dict_hn = {}
cl_count_wrkloc_list = []

# For reading from NP download folder in which all rtr configs are present in sub folders
for index, row in splunk_data.iterrows():
#for hn in splunk_data['Device Name']:
	#print (index, row)
	#print (row["Device Name"])
	hn = row['Device Name']
	if hn == '' or pd.isnull(hn):
		print (" -----------------------------------------------------------------\n")
		print (" Parsing config for Hostname {}\n".format(hn))
		print (" Found null name - skipping")
		continue
	hostcount = hostcount+1	
	
	cdp_file = ""
	wan_neighbors = []
	wan_nei_ip = []
	site_name = ""
	region = ""
	dict_op_hn = {}
	source = ""
	print (" -----------------------------------------------------------------\n")
	print (" Parsing config for Hostname {}\n".format(hn))
	
	#get device details from splunk download
	
	#['Cluster','Country','Work Location','Work Location Code','CMSP - Customer Name','CMSP - Region','CMSP - Site','Device Onboarding Date','Device ID','Asset ID','Device IP','Device Name & Domain','Device Category','Model','Serial','IOS','Device State','Impact','Device Name','Region:Site']

	dict_op_hn["Cluster"] = row['Cluster']
	dict_op_hn["Country"] = row['Country']
	dict_op_hn["Work Location"] = row['Work Location']
	dict_op_hn["Work Location Code"] = row['Work Location Code']
	
	dict_op_hn["CMSP - Site"] = row["CMSP - Site"]
	dict_op_hn["CMSP - Region"] = row['CMSP - Region']
	dict_op_hn["Device IP"] = row['Device IP']
	dict_op_hn["Device Name & Domain"] = row['Device Name & Domain']
	
	if not dict_op_hn["Cluster"] or not dict_op_hn["Country"] or not dict_op_hn["Work Location"]:
		print ("Cluster, country or work location not specified for device ", hn,"\n", row)
		continue
	else:
		cl_count_wrkloc_list.append(str(dict_op_hn["Cluster"])+":"+str(dict_op_hn["Country"])+":"+str(dict_op_hn["Work Location"]))
	neighbors = list(splunk_data[splunk_data['Work Location'] == dict_op_hn["Work Location"]]['Device Name'].items())
	if neighbors:
		for i in neighbors:
			if i[1] != hn:
				wan_neighbors.append(i[1])
		dict_op_hn["wan_neighbors"] = wan_neighbors
		print ("Found WAN neighbors exist for device ", hn)
	else:
		print ("No WAN neighbors exist for device ", hn)
	nei_ip = list(splunk_data[splunk_data['Work Location'] == dict_op_hn["Work Location"]]['Device IP'].items())
	if nei_ip:
		for i in nei_ip:
			if i[1] != dict_op_hn["Device IP"]:
				wan_nei_ip.append(i[1])
		dict_op_hn["wan_nei_ip"] = wan_nei_ip
		print ("Found WAN neighbor IPs exist for device ", hn)
	else:
		print ("No WAN neihgbor IPs exist for device ", hn)
		
	#check where the device data is found 
	if hn in json_dict.keys():#Choice 1: Coral JSON files
		for key in json_dict[hn].keys():
			dict_op_hn[key] = json_dict[hn][key]
		source = "CORAL task download"
		#print (dict_op_hn.keys())
	
	if hn in np_dict.keys():   #Choice 2: NP download
		for key in np_dict.keys():
			dict_op_hn[key] = np_dict[hn][key]
		source = "NP inventory download"
		#print (dict_op_hn.keys())
	
	if hn in dict_op.keys(): #Choice 3: network data download
		for key in dict_op[hn].keys():
			dict_op_hn[key] = dict_op[hn][key]
		source = "show commands from network"
		#print (dict_op_hn.keys())
		
	else:
		print ("Device not found from CORAL, NP or direct device download ", hn)
		source = "Not found"
		continue

	if "show ip route" in dict_op_hn.keys():
		print ("show ip route parsing for ",hn)
		dict_op_hn["sir_parsed"] = do_parse_dict(dict_op_hn["show ip route"],hn)
	
	
	#print (dict_op_hn["show run"])
	print ("Source: ", source)
	
	dict_op_hn = rtr_dict(hn,wb,op_dir,masterTracker, policy_data, dict_op_hn, source, ws_summ , ws,ws_int,ws_rtng,ws_plcy,ws_cdp,ws_wan,ws_dhcp,ws_sdwan,wan_nei_ip,ws_lc,ws_sdwan_rtr,ws_sdwan_site)
	dict_all_hn[hn] = dict_op_hn

timestamp = datetime.now().strftime('%Y-%m-%d-%H_%M')
filnam = "UL_devices" + timestamp + ".xlsx"
wb.save(os.path.join(op_dir,filnam))

#show ip rote comparison

op_routes_dir = os.path.join(op_dir,"Route-summary"+timestamp)
if not os.path.exists(op_routes_dir):
	os.makedirs(op_routes_dir)

dfAllHN_HN = pd.DataFrame(dict_all_hn)
dfAllHN = dfAllHN_HN.transpose()

#Show ip route comparison
row_routes = ws_routes.max_row
if row_routes == 1:
	ws_routes.cell(row_routes,1,"Country")
	ws_routes.cell(row_routes,2,"Work Location")
	ws_routes.cell(row_routes,3,"Hostname")
	ws_routes.cell(row_routes,4,"Neighbor")
	ws_routes.cell(row_routes,5,"Show ip route matches")
	ws_routes.cell(row_routes,6,"Only in device")
	ws_routes.cell(row_routes,7,"Only in neighbor")
	ws_routes.cell(row_routes,8,"OSPF DIO metrics in device")
	ws_routes.cell(row_routes,9,"OSPF DIO metrics in neighbor")
	ws_routes.cell(row_routes,10,"OSPF metrics as standard")

site_count = 0
for item in list(set(cl_count_wrkloc_list)):
	print ("cluster country work location: ", item)
	find_loc = re.search("(\S+)\:(\S+):(\S+)",item)
	site_count+=1
	if find_loc:
		cluster = find_loc.group(1)
		country = find_loc.group(2)
		wrk_loc = find_loc.group(3)
	else:
		continue
	dev_in_site = dfAllHN[dfAllHN['Country'] == country]
	dev_in_region = dev_in_site[dev_in_site["Work Location"] == wrk_loc]
	nei_num , routes_match = compare_routes_of_wan(dev_in_region.transpose(),op_routes_dir,item)
	#print (" Neighbor number", nei_num)
	#print (" Routes matched ", routes_match)
	print_combo_op(routes_match, ws_routes, country, wrk_loc, nei_num)
		
#write to output
freeze_panes (wb, "C2")
wb.remove(wb['SDWAN Readiness summary'])
wb.save(os.path.join(op_dir,filnam))

#Read_Excel = pd.read_excel(os.path.join(op_dir,filnam),sheet_name='RTR summary', usecols=['Region','Site Name','IP','Hostname','Device Type','Routing in use','Existing Tunnel IP','Ordered net mask','Existing net mask on WAN interfaces','NAT Pool','NAT IP start range','NAT IP end range','NAT Mask','NAT Mask in CIDR','ZBFW','ZBFW list','Filename','IP SLA','Bandwidth of physcial interfaces','Circuit ID seen','WAN interface','Device Type','Image'])
#usecols=['RFP_REF','Hostname','Interface','Bandwidth','IP address','Network Mask','Network Mask in CIDR','Interface type','Circuit Type','Circuit ID','Encapsulation','Port in use','NAT','Standby config','Routing protocol(s)','Description','Contents','

#run_nw_chk(ws_chk,ckt_sheet,Read_Excel)

#row_chk = ws_chk.max_row

#freeze_panes (wb1,"B1")
#wb1.remove(wb1['Sheet'])
#filnam = "UL_NW_ckt_chk" + datetime.now().strftime('%Y-%m-%d-%H_%M') + ".xlsx"
#wb1.save(os.path.join(op_dir,filnam))
#print ("CkT details given for :"+str(len(ckt_sheet)))
print ("Configs parsed : "+ str(hostcount) + "\n")
print ("Output at : " + filnam + "\n")
print ('\a')
print ("---- Completed in {} seconds ---- ".format("{:.3f}".format(time.time() - start_time)))