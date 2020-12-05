import pandas as pd
from openpyxl import Workbook
import os,re
from datetime import datetime
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
import time

def readSheetCols(filename,sheetname,cols,hdr):
	reqdData = pd.read_excel(filename, sheet_name=sheetname, usecols=cols, header = hdr)
	##print(reqdData.head())
	return reqdData

#writing with highlight
def write_cell(ws,row,col,value):
	ws.cell(row,col,value)
	if value == "WAN":
		ws[colnum_string(col) + str(row)].fill = redFill
	if value == "" :
		ws[colnum_string(col) + str(row)].fill = redFill

start_time = time.time()
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=15)
bd = Side(style='thick', color="111111")
highlight.border = Border(bottom=bd) #left=bd, top=bd, right=bd, 

redFill = PatternFill(start_color='FFEE08',   #FFFF0000 - red
				   end_color='FFEE08',
				   fill_type='solid')

filename = "./Script_input/C&R/07aug/Audit_Summary_408.xlsm"	
CR_mapping = "./Script_input/C&R/CR to bucket mapping.xlsx"
  
ViolSheet = readSheetCols(filename,'All_Violations',['Host Name','IP Address','Policy Name','Rule Name','Violation Description','Severity'],4)
PassSheet = readSheetCols(filename,'All_Passes',['Host Name','IP Address','Policy Name','Rule Name','Pass Description'],4)
CR_map    = readSheetCols(CR_mapping,'Mapping',['Policy Name','Rule Name','Bucket'],0)
InvSheet  = readSheetCols(filename,'Inventory',['Host Name','IP Address','Product Model','OS Version'],4)

#print ("Viol sheet", ViolSheet)
#print ("Pass sheet", PassSheet)
#print ("CR_map", CR_map)
#print ("Inv sheet", InvSheet)

rules_for_no_sh_run = ["No Running config present","No running Config Present to Audit","No Running Config Present","NO Configs","NO Config","Running config not available to audit"]

op_dir = os.path.join(os.getcwd(),"Script_output")
op_file = "CR_summary-"+ datetime.now().strftime('%Y-%m-%d-%H_%M') + ".xlsx"

wb = Workbook()
ws_summ = wb.create_sheet("CR summary")
wb.remove(wb['Sheet'])
row_summ = ws_summ.max_row


if row_summ == 1:
	ws_summ.cell(row_summ,1,"Hostname")
	ws_summ.cell(row_summ,2,"IP")
	ws_summ.cell(row_summ,3,'Product Model')
	ws_summ.cell(row_summ,4,'OS Version')
	col = 5
	col_dict = {}
	plcy_col_dict = {}
	for key in CR_map["Bucket"].unique():
		ws_summ.cell(row_summ,col,key)
		col_dict[key] = col
		col+=1
		ws_det = wb.create_sheet(key)
		row_det = ws_det.max_row
		ws_det.cell(row_det,1,"Hostname")
		ws_det.cell(row_det,2,"IP")
		ws_det.cell(row_det,3,'Product Model')
		ws_det.cell(row_det,4,'OS Version')
		row_det = ws_det.max_row
		plcy_col_dict[key] = {}
		p_col = 5
		policy_names = CR_map[CR_map["Bucket"] == key]
		print ("All policy names", str(policy_names["Policy Name"]))
		for policy_name in policy_names["Policy Name"].unique():
			plcy_col_dict[key][policy_name] = {}
			print (policy_name)
			for rule_name in policy_names[policy_names["Policy Name"]==policy_name]["Rule Name"].unique():
				ws_det.cell(row_det,p_col,key)
				ws_det.cell(row_det+1,p_col,policy_name)
				ws_det.cell(row_det+2,p_col,rule_name)
				plcy_col_dict[key][policy_name][rule_name] = p_col
				p_col+=1


#add bucket to df for pass and viol
PassBucket = pd.merge(PassSheet, CR_map,  how='left', left_on=['Policy Name','Rule Name'], right_on = ['Policy Name','Rule Name'])
print ("Pass bucket: ", PassBucket)
ViolBucket = pd.merge(ViolSheet, CR_map,  how='left', left_on=['Policy Name','Rule Name'], right_on = ['Policy Name','Rule Name'])
print ("Viol bucket: ", ViolBucket)



#calc and print bucket summary
for rtr in InvSheet["Host Name"]:
	print ("Creating summary for rtr", rtr)
	row_summ+=1
	write_cell(ws_summ, row_summ, 1, rtr)
	write_cell(ws_summ, row_summ, 2, InvSheet[InvSheet["Host Name"] == rtr]["IP Address"].item())
	write_cell(ws_summ, row_summ, 3, InvSheet[InvSheet["Host Name"] == rtr]["Product Model"].item())
	write_cell(ws_summ, row_summ, 4, InvSheet[InvSheet["Host Name"] == rtr]["OS Version"].item())
	for key in col_dict.keys():
		rtr_in_pass = PassBucket[PassBucket['Host Name']== rtr]
		pass_key_match = rtr_in_pass[rtr_in_pass['Bucket'] == key]
		rtr_in_viol = ViolBucket[ViolBucket['Host Name']== rtr]
		viol_key_match = rtr_in_viol[rtr_in_viol['Bucket'] == key]
		pass_row,p_col = pass_key_match.shape
		viol_row,v_col = viol_key_match.shape
		status = "Fail"
		if pass_row > 0 and viol_row >0 :
			status = "Fail"
		elif viol_row >0:
			status = "Fail"
		elif pass_row >0:
			status = "Pass"
		write_cell(ws_summ, row_summ, col_dict[key], status)
		#print ("key", key, " Status: ", status,"\n")
		ws_bucket = wb[key]
		bucket_row = ws_bucket.max_row+1
		write_cell(ws_bucket, bucket_row, 1, rtr)
		write_cell(ws_bucket, bucket_row, 2, InvSheet[InvSheet["Host Name"] == rtr]["IP Address"].item())
		write_cell(ws_bucket, bucket_row, 3, InvSheet[InvSheet["Host Name"] == rtr]["Product Model"].item())
		write_cell(ws_bucket, bucket_row, 4, InvSheet[InvSheet["Host Name"] == rtr]["OS Version"].item())
		for policy_name in plcy_col_dict[key].keys():
			for rule_name in plcy_col_dict[key][policy_name].keys():
				#print ("Rule name", rule_name, " Policy name " , policy_name,"Bucket", key, " Hostname",rtr)
				rtr_in_pass = PassBucket[PassBucket['Host Name']== rtr]
				pass_key_match = rtr_in_pass[rtr_in_pass['Bucket'] == key]
				plcy_pass_match = pass_key_match[pass_key_match['Policy Name'] == policy_name]
				rule_pass_match = plcy_pass_match[plcy_pass_match['Rule Name'] == rule_name]
				#print ("Pass:", rule_pass_match[["Policy Name","Rule Name"]])
				
				rtr_in_viol = ViolBucket[ViolBucket['Host Name']== rtr]
				viol_key_match = rtr_in_viol[rtr_in_viol['Bucket'] == key]
				plcy_viol_match = viol_key_match[viol_key_match['Policy Name'] == policy_name]
				rule_viol_match = plcy_viol_match[plcy_viol_match['Rule Name'] == rule_name]
				#print ("Viol:", rule_viol_match[["Policy Name","Rule Name"]])
				
				pass_row,p_col = rule_pass_match.shape
				viol_row,v_col = rule_viol_match.shape
				
				#print ("pass row: ", pass_row, "viol_row: ", viol_row)
				status = "Fail"
				if pass_row > 0 and viol_row >0 :
					status = "Fail"
				elif viol_row >0:
					status = "Fail"
				elif pass_row >0:
					status = "Pass"
				#print ("Status:" , status)
				write_cell(ws_bucket, bucket_row, plcy_col_dict[key][policy_name][rule_name], status)
				#print ("key", key, "Policy name",policy_name, "Rule Name", rule_name, " Status: ", status,"\n")

		

#for rtr in InvSheet["Host Name"]:
#	print ("Creating detailed policy summary for rtr", rtr)
#	for key in wb.sheetnames:
#		if not "CR summary" in key:
#			ws_bucket = wb[key]
#			bucket_row = ws_bucket.max_row+1
#			write_cell(ws_bucket, bucket_row, 1, rtr)
#			write_cell(ws_bucket, bucket_row, 2, InvSheet[InvSheet["Host Name"] == rtr]["IP Address"].item())
#			write_cell(ws_bucket, bucket_row, 3, InvSheet[InvSheet["Host Name"] == rtr]["Product Model"].item())
#			write_cell(ws_bucket, bucket_row, 4, InvSheet[InvSheet["Host Name"] == rtr]["OS Version"].item())
#			for policy_name in plcy_col_dict[key].keys():
#				for rule_name in plcy_col_dict[key][policy_name].keys():
#					print ("Rule name", rule_name, " Policy name " , policy_name,"Bucket", key, " Hostname",rtr)
#					rtr_in_pass = PassBucket[PassBucket['Host Name']== rtr]
#					pass_key_match = rtr_in_pass[rtr_in_pass['Bucket'] == key]
#					plcy_pass_match = pass_key_match[pass_key_match['Policy Name'] == policy_name]
#					rule_pass_match = plcy_pass_match[plcy_pass_match['Rule Name'] == rule_name]
#					print ("Pass:", rule_pass_match)
#					
#					rtr_in_viol = ViolBucket[ViolBucket['Host Name']== rtr]
#					viol_key_match = rtr_in_viol[rtr_in_viol['Bucket'] == key]
#					plcy_viol_match = viol_key_match[viol_key_match['Policy Name'] == policy_name]
#					rule_viol_match = plcy_viol_match[plcy_viol_match['Rule Name'] == rule_name]
#					print ("Viol:", rule_viol_match)
#					
#					pass_row,p_col = rule_pass_match.shape
#					viol_row,v_col = rule_viol_match.shape
#					status = "Fail"
#					if pass_row > 0 and viol_row >0 :
#						status = "Fail"
#					elif viol_row >0:
#						status = "Fail"
#					elif pass_row >0:
#						status = "Pass"
#					print ("Status:" , status)
#					write_cell(ws_bucket, bucket_row, plcy_col_dict[key][policy_name][rule_name], status)
#					#print ("key", key, "Policy name",policy_name, "Rule Name", rule_name, " Status: ", status,"\n")



wb.save(os.path.join(op_dir,op_file))

print ('\a')
print ("---- Completed in {} seconds ---- ".format("{:.3f}".format(time.time() - start_time)))